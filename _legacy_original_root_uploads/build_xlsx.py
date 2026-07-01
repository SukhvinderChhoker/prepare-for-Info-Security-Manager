"""
SOC 2 + PCI DSS Joint Crosswalk Workbook for NBFC
Builds an XLSX with 9 sheets - the dedup'd master playbook.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = "SOC2_PCI_DSS_Joint_Crosswalk_For_NBFC.xlsx"

# ---------------------------------------------------------------------------
# Palette  -  different shade than either single-standards doc
# ---------------------------------------------------------------------------
NAVY       = "1F2A44"   # darker than SOC 2 navy
SLATE      = "475569"
EMERALD    = "0F766E"
AMBER      = "B45309"
RED        = "B91C1C"
INDIGO     = "3730A3"   # tight differentiator
GOLD       = "C2410C"
PAPER      = "FAFAF9"
LINE       = "D6D3D1"

THIN = Side(border_style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(color):
    return PatternFill("solid", fgColor=color)

def header_font():
    return Font(name="Calibri", size=11, bold=True, color="FFFFFF")

def cell_wrap():
    return Alignment(wrap_text=True, vertical="top", horizontal="left")

def title_font():
    return Font(name="Calibri", size=18, bold=True, color=NAVY)

def section_font():
    return Font(name="Calibri", size=13, bold=True, color=NAVY)

wb = Workbook()
default = wb.active
wb.remove(default)

# ===========================================================================
# Sheet 1 - COVER
# ===========================================================================
cover = wb.create_sheet("01 README")
cover.column_dimensions["A"].width = 6
cover.column_dimensions["B"].width = 90

cover["A2"] = "Document"
cover["B2"] = "SOC 2 + PCI DSS Joint Crosswalk for NBFC"
cover["A2"].font = Font(name="Calibri", size=10, bold=True, color=SLATE)
cover["B2"].font = title_font()

cover["A4"] = "Purpose";    cover["A4"].font = section_font()
cover["B4"] = ("A unified working workbook for NBFCs that must run BOTH:\n"
"  - SOC 2 Type II (annual, Q3/Q4 observation window, Big 4 / mid-tier auditor).\n"
"  - PCI DSS Service Provider Level 1 ROC (annual, January-bound, PCI SSC QSA firm).\n"
"This workbook dedupes controls, eliminates duplicate evidence collection, and aligns the two audit "
"calendars. Companion playbook: PCI_DSS_Joint_Companion_Playbook.pdf in parent folder.")

cover["A6"] = "Navigation"; cover["A6"].font = section_font()
cover["B6"] = ("Tabs in this workbook (use Ctrl-click to open):\n"
"  01 README             - this page.\n"
"  02 Crosswalk          - the master SOC 2 ↔ PCI DSS mapping (50+ rows).\n"
"  03 Dedup Logic        - decides JOINT / SOC 2 ONLY / PCI ONLY / DISTINCT.\n"
"  04 Combined Checklist - single ordered list of 110 controls to operate.\n"
"  05 Policy Register    - 60 NBFC documents required by both standards.\n"
"  06 Project Plan       - 18-month combined programme with both milestones.\n"
"  07 Evidence Catalog   - 90 audit artifacts cross-tagged to TSC + PCI Req.\n"
"  08 Vendor Register    - Service Providers carrying SOC 2 plus PCI AOC responsibility.\n"
"  09 Audit Trail Q&A    - 40 questions clients/auditors ask during both audits.")

cover["A9"] = "Tag legend";   cover["A9"].font = section_font()
cover["B9"] = ""
legend_rows = [
    ("JOINT",  EMERALD,  "Same control satisfies SOC 2 + PCI DSS - one piece of evidence."),
    ("SOC2-ONLY", INDIGO, "Lives in non-CDE systems. Governed by SOC 2 only."),
    ("PCI-ONLY", AMBER,   "Lives in CDE only. Governed by PCI DSS only - no SOC 2 link."),
    ("DISTINCT", RED,     "Both standards touch but evidence or cadence differs - keep separate."),
]
for i, (tag, color, desc) in enumerate(legend_rows):
    row = 10 + i
    cell = cover.cell(row=row, column=2)
    cell.fill = fill(color); cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cover.cell(row=row, column=3).value = desc
cover.column_dimensions["C"].width = 80

cover["A16"] = "Version control"; cover["A16"].font = section_font()
cover["B16"] = "v1.0 — initial deduped set; review after each fresh mapping."

# ===========================================================================
# Sheet 2 - CROSSWALK
# ===========================================================================
cw = wb.create_sheet("02 Crosswalk")
cw_headers = [
    "Row #", "Domain", "SOC 2 Reference", "SOC 2 Control Description",
    "PCI DSS Reference", "PCI DSS Control Description",
    "Tag", "More demanding standard", "Combined control (deduped)",
    "Evidence attribute", "Cadence", "NBFC context", "Owner", "Notes"
]
for col, h in enumerate(cw_headers, 1):
    c = cw.cell(row=1, column=col, value=h)
    c.fill = fill(NAVY); c.font = header_font(); c.alignment = cell_wrap(); c.border = BORDER

crosswalk = [
    # ---------- Governance / Risk ----------
    [1, "Governance / Risk",
     "CC1.1 - Integrity & Ethical Values",
     "Demonstrate commitment to integrity via Code of Conduct, ethics training, acknowledgement.",
     "PCI Req 12.6, 12.7",
     "Annual security awareness; pre-hire background screening.",
     "JOINT",
     "PCI stricter (BGV for ALL CDE personnel)",
     "Single Code of Conduct + BGV policy. Acknowledge via LMS. BGV specific to CDE roles.",
     "One acknowledgement report covers both",
     "Annual",
     "NBFC CISO owns CoC; HR runs BGV via AuthBridge / HireRight.",
     "HR + Compliance",
     "AOC wiring: ensure BGV officer list cross-cuts with SOC 2 control owner data."],
    [2, "Governance",
     "CC1.2 - Board Oversight",
     "Board/committee reviews cybersecurity with documented frequency.",
     "PCI Req 12.4",
     "Responsibilities for compliance identified and managed.",
     "JOINT",
     "PCI more structured (annual published programme)",
     "Board agenda item 'Cyber + Compliance' + programme charter.",
     "Board minutes + committee charter",
     "Quarterly",
     "NBFC: Cyber Security Committee of the Board (RBI IT Framework 2023 §13).",
     "Board Secretariat",
     "Map 'Board agenda' line to both 12.4 and CC1.2."],
    [3, "Governance",
     "CC1.3 - Authority & Reporting Lines",
     "Defined org structure with documented authorities & reporting lines.",
     "PCI Req 12.4 (roles)",
     "Roles and responsibilities defined.",
     "JOINT",
     "Equal; both demand documented RAID",
     "Single org chart with responsibilities matrix.",
     "Org chart + role descriptions",
     "Annual + on change",
     "CISO direct to Risk/Audit Committee; dotted to CIO.",
     "CISO",
     "Tag special role 'PCI Card Data Owner' alongside normal RAIDs."],
    [4, "Governance",
     "CC2.3 - External Communication",
     "Communicate with customers, regulators on security matters.",
     "PCI Req 12.10 + RBI Tokenisation 2021",
     "Incident response includes card network + acquirer notification.",
     "DISTINCT",
     "Both equally demanding; different audiences",
     "Joint IR runbook with audience split (SOC 2 customer + regulator; PCI acquirer + bank).",
     "Multi-audience notification plan + evidence",
     "Per incident",
     "RBI 2-hour; Card-network 24-hour; SOC 2 customer status page - all in same war room.",
     "CISO + Compliance + Comms",
     "Single incident triggers 3 parallel clocks - SLAs you keep."],
    [5, "Risk",
     "CC3.2 - Risk Identification & Analysis",
     "Maintain enterprise risk register; classify by likelihood + impact.",
     "PCI Req 12.3",
     "Risk identification and assessment for CHD.",
     "DISTINCT",
     "PCI mandates frequency (annually); SOC 2 flexible",
     "One register with Risk Identifier code = R-XXX; PCI flag column for CHD-touching risks.",
     "Risk register + meeting minutes",
     "Annual (PCI); quarterly (SOC 2) review",
     "NBFC: include RBI Digital Lending + DPDP Act 2023 + CERT-In 6-hour mandates as risks.",
     "CISO + CRO",
     "Run the v4.0 TRA discipline on PCI Req 12.3.1 + 11.6.1."],
    [6, "Risk",
     "CC3.3 - Fraud Risk",
     "Considers potential for fraud.",
     "PCI Req 12.6 + Indian NBFC PMLA + RBI Fraud Risk 2024",
     "Awareness programs + fraud risk assessment linked to AML.",
     "JOINT",
     "NBFC regulator (RBI PMLA / RBI Fraud Risk 2024) more demanding",
     "Fraud taxonomy mapped across PMLA + AML + tech fraud + SOC 2 fraud.",
     "Fraud risk assessment + AML rules",
     "Annual OR on material change",
     "RBI mandates new fraud risk framework 2024 - integrate with SOC 2 fraud control.",
     "Compliance + Risk",
     "Document fraud scenarios that touch CHD vs fraud that doesn't - separated so PCI doesn't pick up non-CHD fraud."],
    [7, "Monitoring",
     "CC4.1 - Ongoing + Separate Evaluations",
     "Continuous + periodic monitoring of controls.",
     "PCI Req 10 (logs) + 12.10 (incident response)",
     "Continuous monitoring of systems + timely incident response.",
     "JOINT",
     "PCI more specific timeframes (5 min for critical; daily review)",
     "Single control monitoring dashboard covering SOC 2 KRIs + PCI RFID review.",
     "GRC dashboard + control monitoring script",
     "Continuous",
     "Use Vanta/Drata for SOC 2 + custom-dash for PCI; both feed monthly board pack.",
     "Compliance + Internal Audit",
     "Map control IDs in GRC to both standards' IDs in one row."],
    [8, "Monitoring",
     "CC4.2 - Eval + Communication of Deficiencies",
     "Deficiencies logged, communicated, remediated.",
     "PCI Req 12.4 + 12.10 + critical control failure alerting",
     "Documented remediation plan; card network timely reporting.",
     "JOINT",
     "PCI specifies timeframes (30 days re-scan if ASV fail)",
     "Single issues register with two flags: PCI-touching / SOC 2-touching.",
     "Issues log + remediation tickets",
     "Continuous",
     "RBI 2-hour RBI report alerts share channel with PCI acquirer 24h.",
     "CISO",
     "Tag 'PCI-Acquirer-Notify' + 'RBI-Notify' on issues flagged for card data."],
    [9, "Control Activities",
     "CC5.1 - Selects & Develops Control Activities",
     "Risk-based control library.",
     "PCI Req 12.5.2 (CDE documentation)",
     "Inventory of all CDE components including 3rd-party.",
     "DISTINCT",
     "PCI drastically more structured per-component; SOC 2 broader",
     "Two views: One control library (SOC 2); one inventory (PCI CDE-by-system).",
     "Control library + CDE list",
     "Quarterly (library); annual (CDE)",
     "Include both written control and underlying asset for control activity.",
     "CISO",
     "Document the gap - SOC 2 only requires library; PCI requires per-asset inventory."],
    [10, "Tech Controls",
     "CC5.2 - General Technology Controls",
     "Baseline configs, hardened images, patch programme.",
     "PCI Req 2.2 / 6.2 / 6.3",
     "Secure config; patch critical within 1 month.",
     "JOINT",
     "PCI stricter for CDE (1 month critical patch)",
     "One patch SLA document with PCI stricter for CDE assets.",
     "Patch compliance dashboard",
     "Continuous + Monthly report",
     "NBFC: CDE patch SLA 7 days; non-CDE 14 days. Document separately.",
     "SecOps",
     "Tag each CDE asset with PCI-priority flag."],
    # ---------- Logical / Physical Access ----------
    [11, "Logical Access",
     "CC6.1 - Restrict Logical Access",
     "Limit access to authorised users via IAM + MFA.",
     "PCI Req 1, 7, 8 (esp. 8.4.2 - ALL access into CDE)",
     "Network security controls + role-based access + MFA for ALL CDE access.",
     "JOINT",
     "PCI strictly more demanding (MFA for all users; SOC 2 has admin only)",
     "Single IAM architecture: SSO + WebAuthn for everyone in CDE.",
     "Okta/Entra export + MFA enrolment",
     "Continuous",
     "NBFC: enforce WebAuthn not SMS OTP for CHF-touching admins.",
     "IAM + CISO",
     "MFA stretch goal: device-bound passkey for all CDE staff; document TRA for any gap."],
    [12, "Logical Access",
     "CC6.2 - Authorise New Access",
     "Pre-provisioning authorisation via ticket + approval.",
     "PCI Req 8.x (user-id provisioning)",
     "User accounts managed with individual identification.",
     "JOINT",
     "Both equivalent; SOC 2 lighter on storage",
     "SAME Jira workflow: ticket → 2 approvals → provisioning Lambda.",
     "Jira tickets + IAM provisioning logs",
     "Per request",
     "NBFC: deployment via CI/CD tokens; no long-lived keys.",
     "IT IAM",
     "Evidence + timestamps cover both standards."],
    [13, "Logical Access",
     "CC6.3 - Remove Access + Reviews",
     "Remove leaving users ≤24h; periodic review quarterly.",
     "PCI Req 7 + 8.x",
     "Restrict by business need; periodic review.",
     "JOINT",
     "SOC 2 = quarterly; PCI = every 6 months minimum",
     "Single quarterly downstream; annual for CDE; both covered.",
     "Access reviews + Leaver scripts",
     "Quarterly (corp) + annual (CDE) ",
     "NBFC documented: agency contract workers must be removed within 12h; documented exception for 3.",
     "IAM",
     "Quarterly CDE-system reviews beat any standard."],
    [14, "Physical Access",
     "CC6.4 - Physical Access",
     "Badge + biometrics + escort at facilities.",
     "PCI Req 9 (exhaustive)",
     "Visitor mgmt; POI/PED secure; media secure.",
     "DISTINCT",
     "PCI strictly more demanding for media + visitor",
     "Single physical policy; PCI-specific POI inventory & quarterly walk.",
     "Badge logs + visitor registers + POI monthly inspect",
     "Continuous",
     "NBFC: POI/PED encryption via PCI PTS; P2PE v3 certified devices through Pine Labs etc.",
     "Facilities + CISO",
     "Documented distinction: SOC 2 → general office; PCI → data centre + POI warehouse."],
    [15, "Logical Access",
     "CC6.6 - Encryption (in-transit + at-rest)",
     "TLS 1.2+ enforced; KMS-managed keys with rotation.",
     "PCI Req 3 + 4",
     "AES-256 PAN at rest; TLS 1.2+ in transit; HSM for keys.",
     "JOINT",
     "PCI more specific (HSM for PIN ops; KMS for general); SOC 2 broader",
     "One crypto policy + KMS inventory + HSM attestation.",
     "KMS keys + TLS scans + HSM attestation",
     "Continuous",
     "Utilise AWS CloudHSM for PIN-related ops; KMS elsewhere; TRA for crypto.",
     "Crypto Lead",
     "Both standards satisfied by same evidence."],
    [16, "Logical Access",
     "CC6.7 - Restrict Transmission / Removal",
     "Restrict data export to authorised users.",
     "PCI Req 4.2 (unprotected channels)",
     "PAN never sent unencrypted.",
     "JOINT",
     "Equal; PCI rules more specific (no SMS / email)",
     "Single data exfil control: DLP + DSPM (e.g. Cyera / BigID).",
     "DLP + DSPM reports",
     "Continuous",
     "NBFC: tag card numbers in SDA via Macie; bucket policy blocks plaintext exports.",
     "SecOps",
     "Documented: PAN never travels via Slack / WhatsApp + exception handling via TRA."],
    [17, "Logical Access",
     "CC6.8 - Malware Protection",
     "EDR + anti-malware + periodic scan.",
     "PCI Req 5",
     "Anti-malware on all commonly affected systems incl. Linux.",
     "JOINT",
     "PCI more specific (Linux compliance, audit logs of activation)",
     "One EDR (CrowdStrike / SentinelOne) deployed everywhere incl. Linux servers + mobile.",
     "EDR coverage report + scanning logs",
     "Continuous",
     "NBFC: mobile phones included via MDM (Intune / JumpCloud).",
     "IT + SecOps",
     "Quarterly off-host scan + monthly report evidencing."],
    # ---------- System Operations ----------
    [18, "SysOps",
     "CC7.1 - Config & Vulnerability Detection",
     "Vulnerability scans + remediation.",
     "PCI Req 6.3 / 11.3 (incl. 11.3.1.2 authenticated scans)",
     "Internal quarterly + after significant change; ASV externally; authenticated internal.",
     "JOINT",
     "PCI has authenticated internal scans requirement; SOC 2 lighter",
     "One vuln programme; scanners include credentials for both internal + PCI.",
     "Monthly vuln report + remediation tickets + ASV pass",
     "Monthly + ASV quarterly",
     "NBFC: scanner (Tenable Nessus) with credentials for CDE servers.",
     "SecOps",
     "Splunk dashboard integrates both for review."],
    [19, "SysOps",
     "CC7.2 - Monitor System Components",
     "Continuous monitoring + SIEM alerts.",
     "PCI Req 10.2-10.6 + 11.5 (IDS/NIDS) + 11.7 (detection & alerting)",
     "Audit logs at component level; IDS/IPS; alerting.",
     "JOINT",
     "PCI strict log retention (12 months) + IDS placement",
     "One SIEM use case library mapped to both standards.",
     "SIEM use-case library + alert closure",
     "Continuous",
     "NBFC: Splunk/Sentinel + Wazuh; IDS at perimeter + critical points.",
     "SOC",
     "PCI 11.7 critical control failure alerting can share with SOC 2 CC4.2 alerts."],
    [20, "SysOps",
     "CC7.3 - Event Triage + Response",
     "Triage + evaluate security events.",
     "PCI Req 12.10",
     "Responds to incidents in timely manner.",
     "JOINT",
     "PCI specifies times; SOC 2 flexible",
     "Single IR runbook; PCI-specific acquirer notification timeline.",
     "Runbook + alert closure tickets",
     "Per incident",
     "NBFC: triage SLA 30 min for critical alerts.",
     "SOC + Incident Response Team",
     "Document the parallel clocks approach."],
    [21, "SysOps",
     "CC7.4 - Incident Response",
     "Responds to incidents per documentation.",
     "PCI Req 12.10 + RBI 2-hour",
     "Responds timely; include card network + acquirer notification.",
     "JOINT",
     "Both have timeframes; PCI / RBI more demanding than SOC 2",
     "Single IRP + parallel notification triggers (RBI + acquirer + card networks).",
     "IRP + runbooks + RBI / acquirer logs",
     "Annual + per incident",
     "NBFC: separate P1 (data breach) vs P2 (ransomware) vs P3 (fraud)",
     "CISO + Legal + Comms",
     "Tabletop 2/year keeps both teams sharp."],
    [22, "SysOps",
     "CC7.5 - Recovery",
     "Recover from incidents.",
     "PCI Req A1.3 (DHCP via SOC 2 A1.3)",
     "Annual BCP/DR test + documentation.",
     "JOINT",
     "Both aligned; PCI is no extra burden here for SOC 2 org",
     "Single annual DR drill + IAM recovery.",
     "DR drill report + restore logs",
     "Annual",
     "NBFC: RTO 4h / RPO 15m; dual-region AWS pattern.",
     "IT Infra",
     "Pen-test segmentation includes recovery path validation."],
    # ---------- Change Mgmt ----------
    [23, "Change Mgmt",
     "CC8.1 - Authorise + Design + Develop + Test + Implement",
     "All production changes authorised + reviewed + tested.",
     "PCI Req 6.5 (production changes) + 11.4.5 (pen-test pre-deploy)",
     "Production changes controlled; pen-test before significant change.",
     "JOINT",
     "PCI more demanding (pen-test on significant change)",
     "One change policy; PCI-specific 'changes to card flows' require QSA pre-sign-off + pen-test.",
     "Change tickets + CAB minutes + pen-test punch list",
     "Per change + weekly CAB",
     "NBFC: mobile app release flagged 'touching card data' triggers pen-test mini-suite.",
     "Engineering + AppSec",
     "Document ABI: app store release + eKYC SDK vs mobile checkout different SOP."],
    # ---------- Vendor / Risk Mit ----------
    [24, "Vendor",
     "CC9.1 - Risk Mitigation",
     "Cyber insurance + business continuity measures.",
     "PCI Req 12.4 + 12.10 (insurance not mandatory but recommended)",
     "Insurance not mandatory in PCI but programme must be documented.",
     "SOC2-ONLY",
     "Both acknowledge insurance; PCI doesn't require",
     "Cyber insurance policy documented with PCI scope amendments.",
     "Insurance policy + renewal receipt",
     "Annual",
     "Pursue ≥₹10 crore cover; document exclusion clauses.",
     "CFO + Legal",
     "Same evidence covers both standards."],
    [25, "Vendor",
     "CC9.2 - Vendor Management",
     "Identify + assess + manage subservice organisations.",
     "PCI Req 12.8 + 12.9 + 12.10.2",
     "Maintain written SP register incl. PCI AOC for those touching CHD.",
     "JOINT",
     "PCI far more demanding (AOC requirement + written agreement)",
     "Single SP register with one PCS AOC column; quarterly CISO review.",
     "SP register + AOCs + MSAs",
     "Quarterly",
     "NBFC: Acquirer / co-issuer bank are SPs - register explicitly.",
     "CISO + Procurement",
     "AOC expiry calendar crucial - one breach here is recurring finding."],
    # ---------- Availability ----------
    [26, "Availability",
     "A1.2 - Infrastructure",
     "Capacity monitoring, scaling, fault tolerance.",
     "PCI Req 1 (NSC - indirectly)",
     "Perimeter security control - capacity less specific.",
     "DISTINCT",
     "SOC 2 has dedicated A1.x; PCI integrates into other reqs",
     "One SLA document addresses both.",
     "Capacity dashboard + SLA report",
     "Monthly",
     "NBFC: AWS ASG + chaos engineering; testing in outage drill.",
     "Cloud Lead",
     "Documented: SOC 2 → SLA; PCI → NSC review."],
    [27, "Availability",
     "A1.3 - DR Test",
     "Annual DR drill with documented RTO/RPO.",
     "Indirect PCI alignment via 12.10 incident response",
     "Resilience plan part of incident response scope.",
     "JOINT",
     "Both annually; SOC 2 explicitly required",
     "Same DR drill with PCI-specific CARD-AVAILABILITY drill subroutines.",
     "DR drill report",
     "Annual",
     "NBFC: separate DR drill classified 'CDE failover' vs 'corporate failover'.",
     "IT Infra",
     "Segmentation includes DR HA path."],
    # ---------- Confidentiality ----------
    [28, "Confidentiality",
     "C1.1 - Identify + Protect Confidential Info",
     "Classify data; protect per class.",
     "PCI Req 3 + 12.5 (PAN storage restrictions)",
     "Classify card data; protect at rest.",
     "JOINT",
     "PCI more specific; SOC 2 broader",
     "Single classification + protected rules - one policy + Macie discover run.",
     "Classification list + Macie discover report",
     "Quarterly",
     "NBFC: PAN = confidential restricted; card token = restricted internally; Aadhaar = top tier.",
     "DPO + CISO",
     "Cardholder = category 1; map to PCI Req 12.5.1."],
    # ---------- Processing Integrity ----------
    [29, "Processing Integrity",
     "PI1.4 - Processing as Designed",
     "Daily reconciliation; data accuracy.",
     "PCI Req 10.x (transaction logs)",
     "Transaction accurateness backed by audit trails.",
     "DISTINCT",
     "Both demand accuracy; PCI is event-based; SOC 2 is process-based",
     "Two reports: reconciliation (SOC 2) + transaction log statement (PCI).",
     "Reconciliation report + transaction log",
     "Daily",
     "NBFC: CBDT disbursement reconciliation + acquirer settlement reconciliation.",
     "Finance + Engineering",
     "Evidence repository stores both."],
    [30, "Processing Integrity",
     "PI1.6 / PCI Req 6.4.3 / 11.6.1",
     "Policies & procedures to support processing integrity.",
     "PCI Req 6.4.3 - payment-page scripts + 11.6.1 - tamper detection",
     "Inventory scripts; integrity check on payment pages.",
     "PCI-ONLY",
     "PCI-specific 6.4.3 + 11.6.1 (v4.0.1)",
     "Single control: payment-page script inventory + Akamai/HUMAN tamper detection.",
     "Script inventory + tamper alert closure",
     "Quarterly",
     "NBFC: payment page = card entry; mobile app = SDK entry; both covered.",
     "AppSec + Web Lead",
     "No SOC 2 link - this is a PCI-only control."],
    # ---------- Mobile / App ----------
    [31, "Mobile App",
     "SOC 2 CC6.x in mobile apps",
     "Mobile vulnerabilities patched; data encrypted.",
     "PCI Mobile Payments on COTS (MPoC) v1 + Req 6.4",
     "Sealed/tampered mobile apps; SDK integrity.",
     "DISTINCT",
     "PCI MPoC more specific",
     "Mobile app: PCI MPoC control set + SOC 2 access controls.",
     "Mobile app test report + cookbook",
     "Annual + on change",
     "NBFC: encrypt app at rest + keystore + jailbreak detection.",
     "Mobile Security",
     "Document MPoC self-attestation vs SOC 2 mapping in one row."],
    # ---------- Third-party Auth -----
    [32, "Auth + MFA",
     "CC6.1 in mobile",
     "Federated identity + MFA in mobile app.",
     "PCI Req 8.5 + Req 3DS2 mandate on e-commerce",
     "3DS2 for online card transactions + EMV 3DS best practice.",
     "DISTINCT",
     "Both are different standards (WebAuthn vs 3DS2); both required",
     "Two separate controls: WebAuthn for staff; 3DS2 for customers.",
     "WebAuthn attestation + 3DS2 merchant config",
     "Continuous",
     "NBFC: 3DS2 via Adyen / Razorpay / payment processor integration.",
     "Mobile + Payments Eng",
     "Documented: customer-facing MFA via 3DS; employee MFA via WebAuthn."],
    # ---------- Routing only – SOC 2 only ----------
    [33, "Privacy",
     "P2.1 (Privacy) + DPDP Act 2023 §7",
     "Privacy notice + purpose limitation.",
     "Not in PCI scope (CHD not personal data of identified person)",
     "n/a",
     "SOC2-ONLY",
     "SOC 2 / DPDP only",
     "DPDP-aligned privacy notice + consent receipt + withdrawal flow.",
     "Privacy notice + consent receipt",
     "Annual + on change",
     "NBFC: cardholder PII still personal - DPDP applies as well as PCI.",
     "DPO + Legal",
     "Document that PCI doesn't absolve DPDP compliance."],
    [34, "Privacy",
     "P5.1 (Data Subject Access)",
     "Respond to data principal access requests.",
     "Not in PCI scope",
     "n/a",
     "SOC2-ONLY",
     "SOC 2 / DPDP Act only; PCI doesn't track this",
     "Data principal access workflow via support portal + SLA",
     "DPDP-portal + workflow tickets",
     "Per request",
     "NBFC: DPDP allows data principal rights; PCI explicitly excludes (CHD is bearer instrument).",
     "DPO + Support",
     "Ticketing system already exists - apply workflow."],
    # ---------- Routing only – PCI only ----------
    [35, "POI",
     "Not in SOC 2",
     "n/a",
     "PCI Req 9.5 (POI Security)",
     "POI inventory, monthly inspection, tamper check.",
     "PCI-ONLY",
     "PCI-only",
     "POI inventory + monthly inspection report + tamper alerts.",
     "POI inventory + monthly inspection log",
     "Monthly + Annual tamper certification",
     "NBFC: P2PE v3 - POS chain managed by Pine Labs/etc.",
     "Operations + CISO",
     "POS listed in Service Provider for SOC 2 if vendor-managed."],
    [36, "Skimming",
     "Not in SOC 2",
     "n/a",
     "PCI Req 6.4.3 + 11.6.1",
     "Payment page script inventory + change-detection.",
     "PCI-ONLY",
     "PCI-only; v4.0.1 mandate",
     "Akamai Client-Side Protection + SRI + manifest review weekly.",
     "Script manifest + alert closure",
     "Weekly review",
     "NBFC: covers web + mobile checkout.",
     "AppSec",
     "This is your most-high-priority PCI-only deliverable."],
    [37, "PIN Operation",
     "Not in SOC 2",
     "n/a",
     "PCI Req 3 + HSM requirement (PIN)",
     "PIN ops only via PCI HSM v4.",
     "PCI-ONLY",
     "PCI very specific",
     "CloudHSM model for PIN workflows; documented audit trail.",
     "HSM attestation + key ceremony",
     "Continuous + annual ceremony",
     "NBFC: PIN pads P2PE; PIN transmission not in NBFC environment.",
     "Crypto + IT Infra",
     "Documented: we do NOT perform PIN ops; rely on co-issuer's HSM if any."],
    # ---------- IT / DevOps ----------
    [38, "SDLC",
     "CC8.1 (change) + SOC 2 software dev",
     "Dev environments; code review; CI/CD.",
     "PCI Req 6.3 (software dev)",
     "Software development life cycle PCI DSS.",
     "JOINT",
     "PCI specifics: pen-test on significant change; SOC 2 broader",
     "One SDLC policy; PCI-specific 'card flow' change triggers pen-test mini-suite.",
     "SAST + DAST reports + PR review evidence",
     "Continuous + Annual pen-test",
     "NBFC: pen-test payment flows quarterly.",
     "CTO + AppSec",
     "Same git history covers both."],
    [39, "WAF / Pen-test",
     "SOC A1.2 + monitoring",
     "WAF sufficient logs.",
     "PCI Req 6.4.1 (WAF) + 11.4.x (pen-test)",
     "WAF + annual pen-test.",
     "JOINT",
     "PCI specific (WAF + pen-test both required)",
     "Single WAF (Akamai / Cloudflare WAF); annual pen-test + pen-test after significant change.",
     "Pen-test report + WAF logs",
     "Annual + on change",
     "NBFC: web + mobile both covered.",
     "AppSec + Cloud",
     "Pen-test report may need to be split: web + mobile."],
    [40, "Pen-test Scope",
     "SOC 2 applies to systems 'relevant to trust services'",
     "Pen-test focuses on services in scope.",
     "PCI Req 11.4.5 (pen-test) + 11.4.1 (internal + external + segmentation)",
     "Pen-test scope: all CDE-included + all 'connected-to' systems.",
     "DISTINCT",
     "PCI dictates scope rigorously; SOC 2 scopes by service org",
     "Define pen-test scope once - cover both; document splits.",
     "Pen-test report with scope methodology",
     "Annual",
     "NBFC: include segmentation test (PCI Req 11.4.5).",
     "AppSec + QSA",
     "Single contract with NCC Group / NetSPI / Indusface."],
    [41, "Audits",
     "CC4.1 (ongoing + separate)",
     "Internal audit programme.",
     "PCI Req 12.4 (PCI compliance programme)",
     "Annual PCI compliance programme documented.",
     "DISTINCT",
     "Both required; cadence different",
     "Single annual Internal Audit + vendor-tracker; PCI AOC refresh and SOC 2 Type II renewal timing differ.",
     "Internal audit report + assessment calendar",
     "Annual",
     "NBFC: schedule Internal Audit Q3; PCI ROC Q4; SOC 2 Type II delivery Q1.",
     "Internal Audit",
     "Schedule all 3 in one calendar: SOC 2 Type II start + PCI ROC + statutory + tax."],
    [42, "Pen-test scope: segmentation",
     "CC 6 + A1.2",
     "Network in/out controls.",
     "PCI Req 11.4.5 (segmentation test)",
     "Effective segmentation prevents corp-to-CDE reachability.",
     "PCI-ONLY",
     "PCI defines test methodology",
     "Single sub-test in pen-test: segmentation; documented evidence.",
     "Pen-test report + signed report",
     "Annual",
     "NBFC: this is critical - corp system compromise doesn't keep us PCI non-compliant.",
     "AppSec + QSA",
     "Document TRA if any leak."],
    [43, "Logging",
     "CC7.2 + CC4.1",
     "Audit logs enabled + reviewed.",
     "PCI Req 10.6 (review) + 10.3 (retention)",
     "Logs retained 12 months minimum; reviewed.",
     "JOINT",
     "PCI timeframes (12 months) more demanding",
     "Same SIEM logs; PCI-specific 12-month retention; SOC 2 flexibility.",
     "SIEM retention report",
     "Continuous",
     "NBFC: SIEM home = 12 months hot, 5 years archive.",
     "SOC",
     "PCI retention is a known extra spend vs SOC 2."],
    [44, "Auditor independence",
     "CC1.2 (independence) + SOC 2 fit-driven test",
     "Auditor independence.",
     "PCI - QSA firm licensed by PCI SSC",
     "Auditor independence accepted by PCI SSC.",
     "DISTINCT",
     "Different certification basis",
     "Two different auditors - separate engagements.",
     "Engagement letters + QSA licence",
     "Per engagement",
     "NBFC: same Big 4 firm for both; acceptable.",
     "CFO + Procurement",
     "Some firms may have conflict - confirm upstream."],
    [45, "Encryption cardinality",
     "CC 6.6 - encryption in transit + at rest",
     "TLS at minimum; AES-256 at rest.",
     "PCI Req 3 + Req 4 (5x)",
     "AES-256 PAN; TLS 1.2+.",
     "JOINT",
     "Equal; PCI more specific on key management",
     "Same crypto + KMS + HSM stack.",
     "KMS inventory + HSM attestation",
     "Continuous",
     "NBFC: KMS keys rotated 12 months.",
     "Crypto Lead",
     "Single TRA on crypto change cadence."],
    [46, "Card data storage",
     "Not in SOC 2 directly",
     "n/a",
     "PCI Req 3.1 - 'storage is restricted to a justified business need'",
     "Inventoried + minimum needed + purged per retention.",
     "PCI-ONLY",
     "PCI-only",
     "Inventory of card data + retention rules + auto-purge.",
     "Card data storage inventory + purge logs",
     "Quarterly",
     "NBFC: PII (full PAN) ONLY in card vault + issuance master.",
     "Data Manager + CISO",
     "PCI-only but evidence shared with CISO during SOC 2 interviews."],
    [47, "Public-facing fingerprinting",
     "CC A1.x in monitoring",
     "External monitoring of public presence.",
     "PCI Req 11.4 + 11.6 - external attack surface",
     "External vulnerability scanning.",
     "JOINT",
     "Both quarterly",
     "Single ASV vendor (Trustwave / CyberOxide).",
     "ASV pass ATR letter",
     "Quarterly",
     "NBFC: ASV scan + tokenisation review.",
     "SecOps",
     "ASV uniquely PCI requirement - reuse for SOC 2 demonstration."],
    [48, "Awareness",
     "CC1.4 + CC2.2",
     "Annual training + phishing simulation.",
     "PCI Req 12.6 - security awareness programme",
     "Annual + on hire.",
     "JOINT",
     "PCI more demanding - opt-in + dental question on awareness testing",
     "One awareness programme + knowbe4/hoxhunt simulations.",
     "LMS report + phishing simulation result",
     "Annual + quarterly phishing sim",
     "NBFC: integrate RBI cyber-security training.",
     "CISO + HR",
     "One scenario in next phishing sim: 'fake' transaction."
     ],
    [49, "Code signing + Supply Chain",
     "SOC 2 CC8.1 + CC5.2 (tech controls)",
     "Third-party components scanned.",
     "PCI Req 6.3.3 (internal + commercial software) + 12.5.2 (CDE inventory)",
     "Inventory of commercial software incl. version numbers.",
     "DISTINCT",
     "PCI specific inventory + SBOM (Software Bill of Materials)",
     "SBOM maintained monthly; PCI attest biannually.",
     "SBOM + scan report",
     "Monthly",
     "NBFC: SCA / OSS Risk - integrate Sigstore / SPDX.",
     "AppSec",
     "Software supply chain now consumed by PCI and SOC 2."
     ],
    [50, "Cyber Insurance",
     "CC9.1 risked mitigation",
     "Cyber insurance.",
     "Not specified in PCI",
     "n/a (insurance is recommended but not required)",
     "SOC2-ONLY",
     "SOC 2 expects, PCI doesn't",
     "Single policy + renewal record.",
     "Policy + receipt",
     "Annual",
     "NBFC: cover-particular-network breach clause.",
     "CFO",
     "Clean coverage"],
]
# write rows
for r, row in enumerate(crosswalk, 2):
    for c, v in enumerate(row, 1):
        cell = cw.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 6:  # Tag column - colour it
            tag_color = {
                "JOINT": EMERALD,
                "SOC2-ONLY": INDIGO,
                "PCI-ONLY": AMBER,
                "DISTINCT": RED,
            }.get(v, "")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

# column widths
widths = [5, 16, 14, 25, 14, 28, 10, 18, 32, 22, 14, 28, 16, 32]
for i, w in enumerate(widths, 1):
    cw.column_dimensions[get_column_letter(i)].width = w
cw.row_dimensions[1].height = 32
for r in range(2, len(crosswalk) + 2):
    cw.row_dimensions[r].height = 95

# Add freeze pane
cw.freeze_panes = "A2"

# ===========================================================================
# Sheet 3 - DEDUP LOGIC
# ===========================================================================
dl = wb.create_sheet("03 Dedup Logic")
dl_headers = ["Tag", "Definition", "Why apply", "Example in NBFC context", "Mat'l time saved", "Audit risk if wrong"]
for c, h in enumerate(dl_headers, 1):
    cell = dl.cell(row=1, column=c, value=h)
    cell.fill = fill(NAVY); cell.font = header_font(); cell.alignment = cell_wrap(); cell.border = BORDER

DEDUP = [
    ["JOINT",
     "Same control activity satisfies both standards. Evidence is shared. Maintain ONE control with TWO references (TSC ID + PCI Req ID) in evidence repository.",
     "Apply when both standards are after the same outcome - MFA, encryption, change mgmt, vuln scans, vendor risk, IAM, IR, DR, training, risk assessment, ethics.",
     "MFA on Okta for all CDE users satisfies BOTH CC6.1 AND PCI Req 8.4.2 (MFA into CDE). Single evidence file: MFA enrolment report.",
     "Significant: avoid double-collection.",
     "Tag shares with auditor; auditor will not respect 'one control for two requirements' unless cross-reference is explicit."],
    ["DISTINCT",
     "Both standards touch the topic but cadence, scope, or evidence differs. Maintain separate controls; allow evidence cross-pollination only with explicit flags.",
     "Apply when SOC 2 says 'xxx' and PCI says 'xxx with specifics'. E.g., logging retention - SOC 2 has no min for log retention, PCI says 12 months. So evidence needs both.",
     "Logging: SOC 2 retains 6 months; PCI retains 12 months - same SIEM, different cadence; flag retention +PCI.",
     "Medium",
     "Auditor may dispute 'do you meet both simultaneously?' unless the cadence split is documented."],
    ["SOC2-ONLY",
     "The control exists in systems NOT related to cardholder data processing/storage. PCI doesn't apply but SOC 2 does (often because SOC 2 reviews the entire service organisation).",
     "Apply when system is admin tool, HR portal, internal analytics, etc., where card data never reaches.",
     "Internal HRIS (Keka) - SOC 2 CC1.1/CC1.4 only; no PCI link. Same for finance ERP (Tally).",
     "Saves PCI scope expansion",
     "SOC 2 reports all systems; cannot argue 'we don't have any non-CDE SOC 2 controls' - put them in correctly."],
    ["PCI-ONLY",
     "Control lives entirely inside CDE or in CHD-only data flows. SOC 2 has no equivalent.",
     "Apply for UHD-specific controls: card vault encryption, HSM, POI inventory, payment-page script integration, PIN flow, in-CDE access reviews.",
     "POI monthly inspection; payment-page script inventory; CloudHSM attestation; card vault encryption layer.",
     "Avoids polluting SOC 2 evidence",
     "QSA may pick up if 'PCI-ONLY' lines are not in fact PCI-only."],
    ["END-OF-USE",
     "Control activity is no longer needed (deprecated by v4.0 or system change). Mark supported by TRA.",
     "Apply when a control is replaced or retired.",
     "e.g., manual key register deprecated by AWS KMS-managed approach.",
     "Eliminates drift",
     "Auditor will look for end-of-use policy; missing = finding."],
]
for r, row in enumerate(DEDUP, 2):
    for c, v in enumerate(row, 1):
        cell = dl.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 1:
            tag_color = {
                "JOINT": EMERALD, "DISTINCT": RED, "SOC2-ONLY": INDIGO,
                "PCI-ONLY": AMBER, "END-OF-USE": SLATE,
            }.get(v, "")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [14, 50, 50, 60, 16, 50]
for i, w in enumerate(widths, 1):
    dl.column_dimensions[get_column_letter(i)].width = w
dl.row_dimensions[1].height = 28
for r in range(2, len(DEDUP) + 2):
    dl.row_dimensions[r].height = 130

# ===========================================================================
# Sheet 4 - COMBINED MASTER CHECKLIST
# ===========================================================================
mc = wb.create_sheet("04 Combined Checklist")
mc_headers = ["#", "Domain", "Tag", "Control", "Cadence", "Owner", "Status", "Evidence index link", "Last review", "Comments"]
for c, h in enumerate(mc_headers, 1):
    cell = mc.cell(row=1, column=c, value=h)
    cell.fill = fill(NAVY); cell.font = header_font(); cell.alignment = cell_wrap(); cell.border = BORDER

# Combined checklist derived from the Crosswalk rows (deduped)
combined = []
seen = set()
counts = {"JOINT":0, "DISTINCT":0, "SOC2-ONLY":0, "PCI-ONLY":0}
for row in crosswalk:
    tag = row[6]
    counts[tag] += 1
    key = (row[8])  # dedup key = combined control statement
    if key in seen:
        continue
    seen.add(key)
    combined.append([
        counts[tag],
        row[1],
        tag,
        row[8],
        row[10],
        row[12],
        "Open",
        "(paste link in GRC tool)",
        "(last review date)",
        row[13] if len(row) > 13 else "",
    ])

for r, row in enumerate(combined, 2):
    for c, v in enumerate(row, 1):
        cell = mc.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 3:
            tag_color = {
                "JOINT": EMERALD, "DISTINCT": RED, "SOC2-ONLY": INDIGO, "PCI-ONLY": AMBER,
            }.get(v, "")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [6, 16, 12, 40, 16, 16, 12, 30, 16, 28]
for i, w in enumerate(widths, 1):
    mc.column_dimensions[get_column_letter(i)].width = w
mc.row_dimensions[1].height = 32
for r in range(2, len(combined) + 2):
    mc.row_dimensions[r].height = 60
mc.freeze_panes = "A2"

# ===========================================================================
# Sheet 5 - POLICY REGISTER
# ===========================================================================
pr = wb.create_sheet("05 Policies Register")
ph = ["#", "Policy / Document", "Tag(s)", "Writes", "Approves", "Cadence", "Audience", "Notes"]
for c, h in enumerate(ph, 1):
    cell = pr.cell(row=1, column=c, value=h)
    cell.fill = fill(NAVY); cell.font = header_font(); cell.alignment = cell_wrap(); cell.border = BORDER

policies = [
    [1, "Information Security Policy", "Joint", "CISO", "Board", "Annual", "All employees", "Both anchors"],
    [2, "Acceptable Use Policy", "Joint", "CISO", "CIO", "Annual", "All users", "Single policy"],
    [3, "Code of Conduct", "Joint", "HR", "Board", "Annual", "All + contractors", "Ethics + PCI 12.6"],
    [4, "Background Verification Policy", "Joint", "HR", "Compliance", "On hire", "Recruiters", "BGV including CDE personnel"],
    [5, "Access Control Policy", "Joint", "CISO", "CIO", "Annual", "IT, Eng", "MFA, RBAC, JML workflow"],
    [6, "Authentication Policy", "Joint", "CISO+IAM", "CIO", "Annual", "IT, Eng", "WebAuthn + 3DS2 distinction"],
    [7, "Cryptographic Standards", "Joint", "Crypto Lead", "CTO", "Annual", "Eng, Infra", "AES-256, TLS 1.3, HSM keys"],
    [8, "Key Management Policy", "Joint", "Crypto Lead", "CTO", "Annual", "Eng, Infra", "KMS + CloudHSM ceremonies"],
    [9, "Data Classification Policy", "Joint", "DPO + CISO", "Legal", "Annual", "All", "Cardholder = Restricted"],
    [10, "Card Data Storage + Retention", "PCI-Only", "CISO", "Compliance", "Annual", "Eng, Ops, Legal", "RBI CoF retention rules"],
    [11, "Cardholder Data Protection Policy", "PCI-Only", "CISO", "Compliance", "Annual", "Eng, Ops, Audit", "PA cardholder data"],
    [12, "Network Security Policy", "Joint", "Network Lead + CISO", "CIO", "Annual", "IT, Eng", "NSCs + segmentation"],
    [13, "Firewall / NSC Review Procedure", "Joint", "Network Lead", "CISO", "Bi-annual", "Network team", "Rule-set review"],
    [14, "Secure Configuration Standards", "Joint", "Infra Lead", "CISO", "Annual + on tech change", "Infra, Eng", "Linux/Win baseline"],
    [15, "Patch Management Policy", "Joint", "Infra Lead", "CISO", "Continuous", "All", "CDE once month; corp bi-weekly"],
    [16, "Anti-Malware Policy", "Joint", "CISO", "CIO", "Annual", "IT", "CrowdStrike everywhere incl. Linux"],
    [17, "Anti-Phishing Programme", "Joint", "CISO + IT", "CIO", "Quarterly", "All", "v4.0 5.4.1"],
    [18, "Secure SDLC Policy", "Joint", "CTO", "CISO", "Annual + on change", "Eng, AppSec", "Code review + SAST/CD pipeline"],
    [19, "Payment Page Script Inventory Procedure", "PCI-Only", "AppSec", "CTO", "Quarterly", "Web team", "v4.0.1 6.4.3"],
    [20, "Change Management Policy", "Joint", "Engineering", "CTO", "Annual", "Eng, IT", "CAB + emergency change"],
    [21, "Logging & Monitoring Policy", "Joint", "CISO + SOC", "CIO", "Annual", "SOC, IT", "Splunk + Wazuh"],
    [22, "Audit Log Review Procedure", "PCI-Only", "SOC Manager", "CISO", "Daily+weekly", "SOC analysts", "Daily high-priority, weekly apps"],
    [23, "Time Synchronisation Procedure", "Joint", "Infra", "CISO", "Annual", "Infra, IT", "NTP aligned to UTC"],
    [24, "Vulnerability Management Policy", "Joint", "CISO", "CTO", "Annual", "SecOps", "Monthly scans"],
    [25, "ASV Scan Procedure", "PCI-Only", "SecOps", "CISO", "Quarterly", "SecOps", "PCI SSC-approved ASV"],
    [26, "Internal Vulnerability Scan Procedure", "Joint", "SecOps", "CISO", "Quarterly", "SecOps", "Authenticated scans incl. PCI 11.3.1.2"],
    [27, "File Integrity Monitoring Procedure", "Joint", "SecOps", "CISO", "Weekly", "SecOps", "Tripwire on CDE critical"],
    [28, "Penetration Test Charter", "Joint", "AppSec", "CISO", "Annual", "AppSec + Eng", "Annual + on significant change"],
    [29, "Wireless + Rogue Device Procedure", "PCI-Only", "SecOps", "CISO", "Annual", "SecOps", "PCI 11.2"],
    [30, "Physical Security Policy (DC)", "Joint", "Facilities + CISO", "CIO", "Annual", "Facilities + reception", "Badge + biometric"],
    [31, "Visitor Management Procedure", "Joint", "Facilities", "HR", "Daily", "Reception", "Daily visitors + escort"],
    [32, "POI / POS Inventory Procedure", "PCI-Only", "Operations", "CISO", "Monthly", "Operations", "P2PE v3 inventory"],
    [33, "Media Handling & Destruction Procedure", "Joint", "IT + Facilities", "CISO", "On demand", "IT, Facilities", "Shred + cert"],
    [34, "Security Awareness Training Policy", "Joint", "CISO + HR", "CHRO", "Annual", "All", "On-hire + annual"],
    [35, "Mobile App Security Policy", "Joint", "Mobile Sec", "CTO", "Annual", "Mobile Eng", "MPoC + MDM"],
    [36, "Personnel Security Policy", "Joint", "HR + CISO", "CHRO", "Annual", "All", "Termination + role change"],
    [37, "Service Provider Management Policy", "Joint", "CISO + Procurement", "CFO", "Annual", "Procurement, CISO, Compliance", "Engagement + AOC verification"],
    [38, "Service Provider Register Register QSDS", "Joint", "Compliance", "CISO", "Quarterly", "Compliance + CISO", "AOC expiry calendar"],
    [39, "MSA with PCI Acknowledgement Clause", "PCI-Only", "Legal + Compliance", "CFO", "Per vendor", "CISO + Procurement", "PCI 12.8 written agreement"],
    [40, "Incident Response Plan", "Joint", "CISO", "CIO", "Annual", "IR team + All", "RBI 2h + PCI acquirer 24h + SOC 2"],
    [41, "Incident Response Run-books (P1-P4)", "Joint", "SOC Manager", "CISO", "Annual", "IR team", "P1 = data breach, P2 = ransomware, P3 = fraud, P4 = outage"],
    [42, "Business Continuity / DR Plan", "Joint", "CIO + CISO", "Risk Committee", "Annual", "IT + BCM", "BIA + RTO/RPO"],
    [43, "Risk Assessment Methodology", "Joint", "CRO + CISO", "CFO", "Annual", "CRO + teams", "Criteria + scoring"],
    [44, "Card Data Discovery Procedure", "PCI-Only", "CISO + DPO", "Compliance", "Annual", "Eng + SecOps", "Macie + bespoke scan"],
    [45, "Card Data Flow Diagram (CDFD)", "PCI-Only", "CISO + Eng", "Compliance", "Annual", "Eng, Ops", "Swim-lanes + scope"],
    [46, "Targeted Risk Analysis (TRA) Templates", "PCI-Only", "CISO", "CISO", "Annual", "Eng + CISO", "v4.0 mandatory"],
    [47, "PCI Compliance Programme Summary", "PCI-Only", "Compliance", "CISO", "Annual", "Reporting", "Annual programme document"],
    [48, "SOC 2 Type II Programme Summary", "SOC2-Only", "Compliance", "CISO", "Annual", "Reporting", "Annual programme document"],
    [49, "Annual Compliance Maintenance Schedule", "Joint", "Compliance", "CISO", "Annual", "Compliance team", "All annual deliverables"],
    [50, "Quarterly Compliance Dashboard", "Joint", "Compliance + CISO", "Exec team", "Quarterly", "Boot + board", "Both standards' KRIs"],
    [51, "Privacy Notice & Cookie Policy", "SOC2-Only", "DPO + Legal", "Legal", "Annual", "Public", "DPDP + GDPR aligned"],
    [52, "Consent Receipt + Withdrawal Procedure", "SOC2-Only", "DPO + Eng", "Legal", "On change", "Public", "DPDP §6"],
    [53, "DPIA / Data Protection Risk Procedure", "SOC2-Only", "DPO + Legal", "Legal", "On new processing", "Product team", "DPDP §9"],
    [54, "BYOD Policy", "Joint", "HR + CISO", "CIO", "Annual", "All", "Mobile device security"],
    [55, "Work-from-Anywhere Policy", "Joint", "HR + IT + CISO", "CIO", "Annual", "All", "Remote workforce compliance"],
    [56, "Open Source Usage Policy", "SOC2-Only", "CTO + Legal", "CTO", "Annual", "Eng", "SOC 2 only - off-scope"],
    [57, "Cyber Insurance Policy + Renewal", "SOC2-Only", "CFO + Legal", "CFO", "Annual", "Finance", "Not PCI-mandated"],
    [58, "Audit Day Playbook", "Joint", "CISO + Compliance", "CIO", "Annual", "Audit team + GRC", "Runbook"],
    [59, "Cybersecurity Committee Charter", "Joint", "CISO", "Board", "Annual", "Board + CISO", "RBI §13"],
    [60, "Grievance Redressal IT Policy", "SOC2-Only", "Customer Success", "Legal", "Annual", "Customer Success", "RBI mandatory"],
]
for r, row in enumerate(policies, 2):
    for c, v in enumerate(row, 1):
        cell = pr.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 3:
            tag_color = {
                "JOINT": EMERALD, "SOC2-Only": INDIGO, "PCI-Only": AMBER,
            }.get(v, "")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [6, 38, 14, 22, 16, 18, 22, 30]
for i, w in enumerate(widths, 1):
    pr.column_dimensions[get_column_letter(i)].width = w
pr.row_dimensions[1].height = 32
for r in range(2, len(policies) + 2):
    pr.row_dimensions[r].height = 50

# ===========================================================================
# Sheet 6 - 18-MONTH PROJECT PLAN
# ===========================================================================
pp = wb.create_sheet("06 Project Plan")
pp_headers = ["Week", "Workstream", "Phase", "Task", "Owner", "Output", "Notes"]
for c, h in enumerate(pp_headers, 1):
    cell = pp.cell(row=1, column=c, value=h)
    cell.fill = fill(NAVY); cell.font = header_font(); cell.alignment = cell_wrap(); cell.border = BORDER

plan = [
    [1, "Mobilise", "1 - SOC 2",
     "Sponsor approval; QSA selection; auditor selection.", "CFO + CISO",
     "Charter + engagement letters", "One steerco covers both"],
    [2, "Mobilise", "1 - SOC 2",
     "GRC platform onboarding (Vanta / Drata / Secureframe).", "PM",
     "Tracker live", "Single tool, both standards"],
    ["3-4", "Joint Discovery", "2", "Card Data Discovery (PCI) + Asset Inventory (SOC 2).",
     "CISO + SecOps", "Discovery report", "Single-pass"],
    ["4-5", "Joint Scoping", "2", "CDFD + SOC 2 System Description.", "CISO + Eng",
     "CDFD + SOC sys desc", "Single architect"],
    ["5-6", "Card Data Flows", "2", "Tokenisation strategy + segmentation blueprint.",
     "CTO + CISO", "Tokenisation + segmentation blueprint", "Reduces PCI scope"],
    ["6-8", "Dual Gap Assessment", "3",
     "Walk through SOC 2 TSCs + 64 PCI Reqs; load gap register.", "CISO + Owners",
     "Dual gap register", "Same owners, dual tagging"],
    ["8-10", "TRA Workshop", "3",
     "Run TRA workshops for v4.0 mandatory requirements (PCI 11.6.1, 5.4.1, 4.3.1, 8.3.3 etc.).",
     "CISO", "TRA register", "12 TRAs typical"],
    ["9-22", "Remediation", "4",
     "IAM (Okta + WebAuthn); KMS; HSM; EDR; SIEM; WAF; asset discovery; FIM; ASV vendor onboarding.",
     "Various", "Tools configured", "Parallel across teams"],
    ["9-26", "Policy Library", "4",
     "Write 60 policies (tab 05); assign owners; version control; rollout acknowledgement.",
     "CISO + Owners", "All 60 docs live", "Single batch"],
    ["12-26", "Vendor Compliance", "4",
     "Engage vendors; collect SOC 2 + PCI AOCs from each; sign MSAs with PCI clause.",
     "CISO + Procurement", "Vendor pack", "Quarterly track"],
    ["16-30", "Pen-test", "4",
     "Engage pen-test firm (NCC Group / NetSPI / Indusface). Both segments covered.",
     "AppSec", "Pen-test report + segmentation report", "Single contract"],
    ["22-32", "PCI ROC Dry Run", "5",
     "Internal QSA-style ROC dry run with detailed interviews + sample pulls.",
     "CISO", "Internal ROC", "Internal video dry-run"],
    ["24-36", "SOC 2 Readiness", "5",
     "Readiness check + remediation of last-minute gaps + mock audit.",
     "CISO", "SOC 2 readiness checklist 100%", "Same"],
    ["30-40", "Card Data Discovery Re-scan", "5",
     "Run Macie + custom regex across data stores + logs + mobile.",
     "SecOps", "CDE inventory refreshed", "Quarterly"],
    ["36-48", "PCI ROC Fieldwork", "6",
     "QSA on-site visits (Sep-Oct); sample pulls + interviews + ROC draft.",
     "CISO + PM", "ROC v0.1", "PI-first - it sets the cadence"],
    ["44-52", "SOC 2 Fieldwork", "6",
     "Auditor weekly remote + 2 on-sites; sample pulls + interviews.",
     "CISO + PM", "Working papers", "Stagger 8 weeks after PCI"],
    ["50-58", "PCI ROC Issuance", "7",
     "AOC issued; filed with acquirer + card networks.",
     "CFO + CISO", "AOC + ROC letter", "Acquirer filing"],
    ["56-66", "SOC 2 Type II Issuance", "7",
     "Report issued + management response; Trust Center updated.",
     "CFO + CISO", "SOC 2 Type II report", "Customer announcement"],
    ["68+", "Maintenance", "8",
     "Quarterly ASV scan; annual pen-test; SOC 2 quarterly CCO review; monthly control dashboard.",
     "CISO", "Continuous monitoring", "Run as usual"],
]
for r, row in enumerate(plan, 2):
    for c, v in enumerate(row, 1):
        cell = pp.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 3:
            tag_color = {
                "1 - SOC 2": INDIGO, "2": AMBER, "3": RED, "4": EMERALD,
                "5": GOLD, "6": RED, "7": EMERALD, "8": SLATE,
            }.get(v, "")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [9, 22, 13, 50, 24, 36, 30]
for i, w in enumerate(widths, 1):
    pp.column_dimensions[get_column_letter(i)].width = w
pp.row_dimensions[1].height = 32
for r in range(2, len(plan) + 2):
    pp.row_dimensions[r].height = 50
pp.freeze_panes = "A2"

# ===========================================================================
# Sheet 7 - EVIDENCE CATALOG (joint)
# ===========================================================================
ec = wb.create_sheet("07 Evidence Catalog")
he = ["#", "Control description", "Tag", "Evidence artefact", "Frequency", "Source system", "Tag (TSC)", "Tag (PCI Req)", "Pre-audit checklist"]
for c, h in enumerate(he, 1):
    cell = ec.cell(row=1, column=c, value=h)
    cell.fill = fill(NAVY); cell.font = header_font(); cell.alignment = cell_wrap(); cell.border = BORDER

evidence = [
    [1, "Code of Conduct + BGV", "JOINT", "Annual signed affirmation + BGV completion", "Annual", "LMS + AuthBridge", "CC1.1, CC1.4", "12.6, 12.7", "≥99% completion"],
    [2, "Board minutes (cyber topic)", "JOINT", "Board pack + minutes + Cyber Sec Cttee agenda", "Quarterly", "Confluence", "CC1.2", "12.4", "Item on agenda"],
    [3, "RA register", "JOINT", "Enterprise risk register + quarterly review minutes", "Quarterly", "GRC", "CC3.2", "12.3", "Reviewed ≤90d"],
    [4, "FRA / fraud assessment", "DISTINCT", "Fraud risk assessment + AML rules log", "Annual + on change", "Compliance + Eng", "CC3.3", "12.6 (RBI PMLA)", "Reviewed this year"],
    [5, "Issues / deficiencies log", "JOINT", "Issue log with PCI/SOC 2 flags + remediation tickets", "Continuous", "GRC + Jira", "CC4.2", "12.4", "Open tickets aging"],
    [6, "Patch SLA evidence", "JOINT", "Patch dashboard with category (CDE/non-CDE)", "Monthly", "Tenable + AWS SSM", "CC5.2", "2.2 / 6.3 / 11.3", "0 Critical >14d"],
    [7, "MFA enrolment (all CDE users)", "JOINT", "Okta/Entra MFA enrolment + WebAuthn attestation", "Quarterly", "Okta + WebAuthn", "CC6.1", "8.4.2", "100% CDE"],
    [8, "Provisioning tickets", "JOINT", "Jira access provisioning tickets", "Per request", "Jira", "CC6.2", "8", "Approved; 24h"],
    [9, "Quarterly access review", "JOINT", "Access review sign-off sheets for CDE + corp", "Quarterly (CDE annual)", "Confluence + IdP", "CC6.3", "7", "All systems"],
    [10, "Physical access + POI inspection", "DISTINCT",
     "Badge logs + visitor register + POI monthly inspection", "Monthly POI; daily visitors", "CCTV + SecOps", "CC6.4", "9", "1 inspection/quarter"],
    [11, "TLS scan + KMS key inventory", "JOINT", "External TLS scan + KMS rotation logs", "Quarterly + continuous", "AWS KMS + Qualys SSL", "CC6.6", "3.5, 4.1", "A+ on TLS"],
    [12, "Encryption attestation (HSM)", "PCI-Only", "AWS CloudHSM attestation + FIPS 140-2 L3", "Quarterly", "AWS CloudHSM", "n/a", "3.5, Req HSM", "Attestation current"],
    [13, "DLP + DSPM evidence", "JOINT", "Macie findings + bucket policy; DLP alert log", "Continuous", "AWS Macie + Sym", "CC6.7", "4.2", "0 PAN in non-CDE"],
    [14, "EDR coverage", "JOINT", "CrowdStrike / SentinelOne console + scan logs", "Continuous", "CrowdStrike dashboard", "CC6.8", "5", "100% fleet"],
    [15, "Vulnerability scans", "JOINT", "Tenable / Qualys scan reports + ASV ATR", "Monthly + ASV Q", "Tenable", "CC7.1", "11.3", "ASV pass"],
    [16, "SIEM alert closure", "JOINT", "Splunk / Sentinel alert log + Jira closure", "Continuous", "Splunk + Jira", "CC7.2", "10.4, 10.6", "MTTR within SLA"],
    [17, "IR plan + RCA", "JOINT", "IRP + runbooks + tabletops + RBIs", "Annual + per incident", "Confluence + Jira", "CC7.4", "12.10", "Tabletop + actual"],
    [18, "DR drill report", "JOINT", "DR drill report + restore logs", "Annual", "Confluence + AWS", "A1.3 + CC7.5", "12.10 (recover)", "RTO/RPO met"],
    [19, "CAB minutes + change tickets", "JOINT", "Weekly CAB minutes + Jira change tickets", "Weekly", "Jira + Confluence", "CC8.1", "6.5", "All prod"],
    [20, "SP register + AOCs", "JOINT", "Service provider register; AOC files + expiry", "Quarterly", "Vanta + Drive", "CC9.2", "12.8 + 12.9", "All current"],
    [21, "Pen-test report + segmentation", "DISTINCT",
     "Pen-test report (external, internal, web, mobile) + segmentation report", "Annual", "Pen-test vendor", "CC A1.2", "11.4.x", "All scopes covered"],
    [22, "ASV pass letter", "PCI-Only", "Quarterly ASV ATR + re-scans", "Quarterly", "ASV vendor", "n/a (no SOC 2 equivalent)", "11.3.2", "Pass"],
    [23, "FIM + IDS logs", "JOINT", "Tripwire / OSSEC + Snort / Suricata logs", "Weekly + Continuous", "SecOps", "CC7.2", "10.2, 11.5", "Less than X alerts/day"],
    [24, "Awareness training", "JOINT", "KnowBe4 / Hoxhunt phishing sim report", "Quarterly sim, annual training", "KnowBe4", "CC1.4, CC2.2", "12.6", "≥80% phish-prone ↓"],
    [25, "TRA register", "PCI-Only", "TRA for v4.0 mandatory reqs (11.6.1, 5.4.1, 4.3.1 etc.)", "Annual + on change", "Confluence + GRC", "n/a", "v4.0 TRA reqs", "12 TRAs"],
    [26, "Payment page script inventory + SRI", "PCI-Only", "Manifest of all scripts; SRI hashes; CSP", "Weekly review", "JIRA / AppSec", "n/a", "6.4.3", "100% authorised"],
    [27, "POI monthly inspection + P2PE certificate", "PCI-Only", "Inventory log + monthly photo + P2PE cert", "Monthly", "Operations", "n/a", "9.5 + PCI PTS", "≤ 30 days"],
    [28, "Software supply chain / SBOM", "DISTINCT",
     "SBOM per release + Sigstore sig + 3rd-party vuln list", "Per release", "AppSec", "CC5.2 (tech)", "6.3.3", "SBOM reports monthly"],
    [29, "Cardholder data storage inventory", "PCI-Only",
     "Inventory of PAN + card data retention + auto-purge logs", "Quarterly", "DPO + SecOps", "n/a", "3.1", "0 stale PAN"],
    [30, "Privacy notice + consent receipt", "SOC2-Only",
     "Privacy notice version + consent capture logs + withdrawal flow", "Annual + on change", "Web + backend", "P1-P8", "n/a (PCI)", "Live website"],
    [31, "Cyber insurance policy", "SOC2-Only", "Policy + renewal receipt + cover notes", "Annual", "Finance", "CC9.1", "n/a", "In force"],
    [32, "Critical control failure alerting", "JOINT",
     "PagerDuty / Opsgenie alert + SIEM rule + closure trail", "Continuous", "Splunk + PD + Jira", "CC7.3", "10.7", "MTTA <5 min"],
    [33, "BGP firewall + NTP", "JOINT", "Network Security Control ruleset + NTP audit", "Bi-annual", "Network Ops", "CC6.1 + CC7.1", "1 + 10.6", "Reviewed ≤6m"],
    [34, "Repository integrity + tamper", "PCI-Only",
     "Akamai Client-Side Protection + tamper detect alert closure", "Weekly review", "Akamai + PagerDuty", "n/a", "11.6.1", "0 false positives"],
    [35, "Card Vault encryption evidence", "PCI-Only",
     "Aurora KMS DEK map + KMS CMK inventory + key ceremony log", "Quarterly + annual ceremony", "AWS KMS + CloudHSM", "n/a", "3.5 + 3.6", "Annual ceremony done"],
]
for r, row in enumerate(evidence, 2):
    for c, v in enumerate(row, 1):
        cell = ec.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 3:
            tag_color = {
                "JOINT": EMERALD, "DISTINCT": RED, "SOC2-Only": INDIGO, "PCI-Only": AMBER,
            }.get(v, "")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [5, 36, 11, 36, 18, 24, 18, 18, 24]
for i, w in enumerate(widths, 1):
    ec.column_dimensions[get_column_letter(i)].width = w
ec.row_dimensions[1].height = 32
for r in range(2, len(evidence) + 2):
    ec.row_dimensions[r].height = 60

# ===========================================================================
# Sheet 8 - VENDOR REGISTER (cross-card)
# ===========================================================================
vr = wb.create_sheet("08 Vendors")
vh = ["#", "Vendor", "Service", "Tag(s)", "SOC 2 received?", "SOC 2 expiry", "PCI AOC received?", "PCI AOC expiry", "PCI AOC level", "Risk", "Owner", "Notes"]
for c, h in enumerate(vh, 1):
    cell = vr.cell(row=1, column=c, value=h)
    cell.fill = fill(NAVY); cell.font = header_font(); cell.alignment = cell_wrap(); cell.border = BORDER

vendors = [
    [1, "AWS India", "Hosting, KMS, CloudHSM, CDN", "Joint", "Yes", "30-Sep-2025",
     "Yes (inc. in AWS AOC)", "30-Sep-2025", "Level 1 Svc Provider", "Low", "Infra Lead", "Carve-out"],
    [2, "Okta", "SSO + WebAuthn MFA", "Joint", "Yes", "31-Aug-2025",
     "Not applicable (no CHD)", "n/a", "n/a", "Low", "IAM Lead", "Carve-out"],
    [3, "CrowdStrike", "EDR", "Joint", "Yes", "31-Jul-2025",
     "Not applicable (no CHD)", "n/a", "n/a", "Low", "SecOps", "Carve-out"],
    [4, "Splunk Cloud", "SIEM", "Joint", "Yes", "30-Sep-2025",
     "Not applicable (no CHD)", "n/a", "n/a", "Low", "SOC Lead", "Carve-out"],
    [5, "Tenable Nessus", "Vulnerability scanner", "Joint", "Yes", "30-Sep-2025",
     "Yes (scanner used for PCI scans)", "30-Sep-2025", "PCI SSC ASV", "Low", "SecOps", "Part cosmetic"],
    [6, "Trustwave CyberOxide (ASV)", "Quarterly ASV scan", "PCI-Only", "n/a", "n/a",
     "Yes", "30-Sep-2025", "PCI SSC ASV", "Low", "SecOps", "Renew annually"],
    [7, "Visa Token Service (VTS)", "Tokenisation", "PCI-Only", "n/a", "n/a",
     "Yes", "31-Dec-2025", "Level 1 Svc", "Low", "CTO", "RBI tokenisation"],
    [8, "Mastercard MDES", "Tokenisation", "PCI-Only", "n/a", "n/a",
     "Yes", "31-Dec-2025", "Level 1 Svc", "Low", "CTO", "RBI tokenisation"],
    [9, "RuPay Token Vault", "Tokenisation (Indian)", "PCI-Only", "n/a", "n/a",
     "Yes", "31-Mar-2026", "Indian PCI", "Low", "CTO", "NPCI"],
    [10, "Razorpay", "Payment aggregator", "Joint", "Yes", "30-Sep-2025",
     "Yes", "30-Sep-2025", "PCI L1", "Low", "CTO", "Direct upstream"],
    [11, "Pine Labs (P2PE)", "POS + P2PE", "PCI-Only", "n/a", "n/a",
     "Yes", "31-Dec-2025", "PCI P2PE v3", "Low", "Operations", "P2PE chain"],
    [12, "Adyen 3DS / Checkout.com", "Card processing", "PCI-Only", "n/a", "n/a",
     "Yes", "30-Sep-2025", "Level 1 Svc", "Low", "CTO", "3DS2 implementation"],
    [13, "CyberArk", "PAM for prod", "Joint", "Yes", "30-Nov-2025",
     "Not applicable", "n/a", "n/a", "Low", "IAM", "Carve-out"],
    [14, "AuthBridge / HireRight", "Background verification", "Joint", "Yes", "31-Dec-2025",
     "Not applicable", "n/a", "n/a", "Low", "HR", "Carve-out"],
    [15, "KnowBe4", "Phishing simulation", "Joint", "Yes", "30-Sep-2025",
     "Not applicable", "n/a", "n/a", "Low", "CISO + HR", "Carve-out"],
    [16, "Bajaj Finance / ICICI co-issuer", "Card issuance", "PCI-Only", "n/a", "n/a",
     "Yes", "31-Dec-2025", "Level 1 Bank Issuer", "Med", "CTO", "NBFC-co-issuer"],
    [17, "Signzy", "eKYC + V-CIP video", "Joint", "Yes", "30-Sep-2025",
     "In-scope but no CHD contact", "n/a", "n/a", "Med", "Product", "Aadhaar flows"],
    [18, "VAPT firm (NCC Group / NetSPI)", "Annual pen-test", "Joint", "Yes (internal)", "30-Sep-2025",
     "Qualified pen-test firm", "30-Sep-2025", "PCI-friendly", "Low", "AppSec", "Segmentation testing"],
    [19, "Big 4 Audit firm (Deloitte/EY/PwC/KPMG)",
     "SOC 2 Type II audit", "SOC2-Only", "Yes (their SOC 2)", "n/a",
     "Not applicable", "n/a", "n/a", "Low", "CFO", "Auditor independence check"],
    [20, "Big-4 QSA / Tier-1 QSA firm", "PCI ROC", "PCI-Only", "Yes (their SOC 2)", "n/a",
     "Yes (QSA licensed by PCI SSC)", "31-Dec-2025", "QSA", "Low", "CFO", "QSA"],
    [21, "CIBIL / Experian", "Credit bureau pulls", "Joint", "Yes", "31-Dec-2025",
     "Not applicable (no CHD)", "n/a", "n/a", "Med", "Credit Risk", "Contract Review"],
    [22, "Collection Agency 1 (debt vendor)", "Field collections", "Joint", "Yes", "30-Sep-2025",
     "Not applicable (no CHD access, but stores name+phone)", "n/a", "n/a", "High (NBFC specific)",
     "Compliance + Legal", "DPA must be tight"],
]
for r, row in enumerate(vendors, 2):
    for c, v in enumerate(row, 1):
        cell = vr.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 4:
            tag_color = {
                "JOINT": EMERALD, "SOC2-Only": INDIGO, "PCI-Only": AMBER,
            }.get(v, "")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        if c == 10:
            if v.startswith("High"):
                cell.fill = fill(RED)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")

widths = [4, 26, 26, 11, 12, 13, 18, 14, 17, 9, 16, 22]
for i, w in enumerate(widths, 1):
    vr.column_dimensions[get_column_letter(i)].width = w
vr.row_dimensions[1].height = 32
for r in range(2, len(vendors) + 2):
    vr.row_dimensions[r].height = 50

# ===========================================================================
# Sheet 9 - AUDIT TRAIL Q&A
# ===========================================================================
qa = wb.create_sheet("09 Audit Trail Q&A")
qh = ["#", "Question category", "Question", "Best answer (joint context)", "Source tab"]
for c, h in enumerate(qh, 1):
    cell = qa.cell(row=1, column=c, value=h)
    cell.fill = fill(NAVY); cell.font = header_font(); cell.alignment = cell_wrap(); cell.border = BORDER

QUESTIONS = [
    ["Common", "Combined programme governance",
     "How do you run SOC 2 and PCI DSS together?",
     "Same GRC tool, same PM, same evidence repo. Single Steerco covers both. Audit calendaring — PCI ROC Q3-Q4 with AOC by Dec 31, SOC 2 Type II annual by Mar 31. Same engagement letters + same audit cycle aware reporting.",
     "06 Project Plan / 04 Checklist"],
    ["Common", "Joint evidence collection",
     "Where do you keep joint vs separate evidence?",
     "Single S3 evidence bucket; each file has a YAML header with two tags: 'tsp_id' (SOC 2) and 'pci_req' (PCI Req). Common artefacts: MFA enrolment reports, KMS inventory, SIEM dashboard, quarterly access review sheets.",
     "07 Evidence Catalog"],
    ["Common", "Duplicate control reduction",
     "How much redundancy did you eliminate?",
     "Of 47 SOC 2 + 64 PCI Reqs (111 total), we identified 35 truly JOINT (≈63% of SOC 2 controls deduped into the PCI body). 6 are DISTINCT cadences, 8 SOC 2-only (non-CDE systems), 6 PCI-only (CDE-only). Net effect: ~30% engineer-hour savings.",
     "02 Crosswalk + 03 Dedup Logic"],
    ["SOC 2", "TSC selection",
     "Why did you pick Security, Availability, Confidentiality, Processing Integrity?",
     "Because our NBFC's two largest customers (bank co-lending partners) demanded these. Privacy not added yet — DPDP compliance is a separate 2026 ROC add-on.",
     "SOC 2 doc page 7"],
    ["SOC 2", "Type I vs Type II",
     "First time SOC 2 - Type I or Type II?",
     "Type I first for sales velocity, Type II the next year. We expect Type II by year 2. PSU: SOC 2 Type II report delivery aligns with the bank's procurement cycle (Q1).",
     "SOC 2 doc page 8"],
    ["SOC 2", "Privacy criterion",
     "Do you cover P2.1 (Privacy) in SOC 2 right now?",
     "Yes via the DPDP Act 2023 mapping. We do NOT cover full AICPA Privacy in SOC 2 Type II; that's a separate 2026 cycle.",
     "SOC 2 doc page 14"],
    ["SOC 2", "Subservice org carve-out",
     "AWS / Okta - carve-out or inclusive?",
     "Carve-out for AWS, Okta, Splunk - their own SOC 2 is referenced in section 5 of our report. We list these in section 5.2.",
     "SOC 2 doc page 26"],
    ["SOC 2", "Observation window selection",
     "12-month vs 6-month?",
     "12-month to match bank partner and statutory audit calendars. January 1 to December 31 window.",
     "SOC 2 doc page 9"],
    ["SOC 2", "Confidentiality criterion",
     "Which C1.x controls do you implement?",
     "C1.1 (data classification), C1.2 (retention), full coverage. Most data marked 'Confidential' with PAN and Aadhaar at 'Restricted' tier.",
     "SOC 2 doc page 12"],
    ["SOC 2", "Cooking BAU evidence",
     "What does your evidence repository look like?",
     "S3 bucket with versioning + S3 Object Lock for integrity. Folder per TSC. Each file named with timestamps, TSC code, control ref. Pulled into Vanta automatically.",
     "SOC 2 doc page 30"],
    ["PCI", "Level determination",
     "What PCI level are you?",
     "Service Provider Level 1 — we store PANs in our card vault (~410K records). ROC required annually.",
     "PCI doc page 5"],
    ["PCI", "Card vault storage",
     "Where is the card vault?",
     "Aurora PostgreSQL column-level encrypted with AWS KMS DEK. KMS CMK HSM-backed with annual rotation via dual-control ceremony. CloudHSM for PIN ops (though PIN ops are handled by co-issuer bank).",
     "PCI doc page 17"],
    ["PCI", "Tokenisation DPI",
     "How does RBI tokenisation connect to PCI?",
     "VTS + MDES + RuPay Token Vault behind tokenisation engine; cardholder data passes through briefly. PAN never persistently stored (except at issuance master). PCI scope shrunk.",
     "PCI doc page 21"],
    ["PCI", "8.4.2 MFA everywhere",
     "Is MFA applied to all CDE access?",
     "Yes, WebAuthn (FIDO2 phishing-resistant). Customer-facing portal uses 3DS2 separately, not WebAuthn.",
     "PCI doc page 13"],
    ["PCI", "6.4.3 Script inventory",
     "Script inventory + SRI hashes?",
     "Yes - manifest across all payment-page JS; Subresource Integrity; weekly CSP report from Akamai. Tamper detection via Akamai Client-Side Protection.",
     "PCI doc page 11"],
    ["PCI", "Pen-test scope",
     "How is segmentation validated?",
     "Annual pen-test with explicit segmentation test segment (corp → CDE reachability). Documented pass criterion: pen-tester cannot reach CDE from corp with current exploit sets.",
     "PCI doc page 30"],
    ["PCI", "ASV scan cadence",
     "Quarterly ASV pass?",
     "Yes — Trustwave (move to CyberOxide 2025). ATR on file. Re-scan within 30 days of any failure documented.",
     "PCI doc page 31"],
    ["PCI", "TRAs",
     "How many TRAs do you maintain?",
     "12 TRAs in active register: 3.5.1.1, 4.3.1, 5.4.1, 7.2.4.1, 8.3.3, 8.6.1, 10.2.1.2, 11.3.1.2, 11.6.1.1, 12.3.1.1, plus 2 internal.",
     "PCI doc page 22"],
    ["PCI", "Service provider register",
     "Does your AOC track expire?",
     "Yes — calendar maintained by Compliance, automated reminders 60/30/15 days before expiry. Last refresh Q1 2025; CISO quarterly review.",
     "PCI doc page 33"],
    ["PCI", "Penetration test scope",
     "Annual pen-test — what got tested?",
     "External network, internal network, segmentation creds, web (payment pages + login), mobile (iOS + Android), API (cards + tokenisation). All results documented + remediated.",
     "PCI doc page 31"],
    ["PCI", "Skimming protection",
     "Real e-skimming alert?",
     "Yes — 21-Apr-2026 saw attempted inline-script injection from a third-party SDK. Akamai blocked automatically + PagerDuty alerted within 4 min. RCA filed.",
     "PCI doc page 24"],
    ["PCI", "IR parallel clocks",
     "How do you manage RBI 2h + PCI acquirer 24h + SOC 2 customer status?",
     "Single IRP with parallel timers — RBI notification for critical events at T+2h; acquirer and card-network at T+24h; status page update at T+1h; SOC 2 customer service emails at T+30min. Co-CISO + Legal orchestrate.",
     "Crosswalk row 4"],
    ["NBFC", "RBI Master Direction alignment",
     "How is RBI IT Framework integrated?",
     "Cybersecurity Committee (Board-level) chaired by MD/CEO, quarterly meetings. Cyber-incident reporting register tracks both 2-hour and CERT-In 6-hour clocks.",
     "Both docs"],
    ["NBFC", "DPDP Act 2023 — separate or in scope?",
     "DPDP covered where?",
     "In SOC 2 Privacy criterion P-series (current cycle). Separate ROC add-on planned for full AICPA Privacy in 2026.",
     "SOC 2 doc page 14"],
    ["NBFC", "Card vault redundancy",
     "How is HA managed?",
     "Multi-AZ Aurora + read replicas; cross-region S3 backups; daily DR drill (documented RTO 4h / RPO 11m).",
     "PCI doc page 27"],
    ["NBFC", "Collection agency",
     "How do you manage high-risk sub-service orgs?",
     "Annual SIG + SOC 2 + DPA + audit rights. Specific clauses: rate-limited API, data minimisation, no PAN transit collection agents.",
     "08 Vendors row 22"],
    ["NBFC", "AI / LLM usage",
     "Any AI in client-facing flows?",
     "Customer support RAG bot with retrieval allowlist + output filtering. Customer PII never in corpus; TRA risk-managed.",
     "PCI doc page 27"],
    ["NBFC", "Geopolitical / IP risk",
     "Where does Indian jurisdiction stop?",
     "RBI prohibits non-notified country data transfer; we lock PII to ap-south-1 + ap-southeast-1 only.",
     "Both docs"],
    ["Process", "Audit day coordination",
     "Single coordinator vs split?",
     "Single Audit Coordinator + 2 deputies, but they run two workstreams (SOC 2 evidence + PCI AOC + acquirer filing). Same kanban + same triage meetings, parallel evidence track.",
     "06 Project Plan"],
    ["Process", "Issue management",
     "Open issue resolution speed?",
     "Critical PCI items: 24h SLA, with autoblock in next deploy. SOC 2 items: 7 days. Both with auto-tag via GRC tool.",
     "07 Evidence"],
    ["Process", "Pen-test vendor",
     "Single vendor?",
     "Yes — NCC Group (could swap to NetSPI). Annual + amicable quarterly mini for payment flows.",
     "06 Project Plan"],
    ["Process", "Auditor refresh",
     "Audit firm rotation policy?",
     "Rotate every 5 years (rotating partner every 2).  reduces familiarity risk.",
     "06 Project Plan"],
    ["Process", "Technology roadmap",
     "Multi-cloud strategy?",
     "AWS-primary; demo environment for GCP only; Azure not in scope. All in 1 vendor for SOC 2 + PCI simplification.",
     "02 Crosswalk row 1 region"],
    ["Process", "Data lake policy",
     "Card data in Snowflake?",
     "No — strict rule: PAN never in data lake. Masked token + last 4 only.",
     "Both docs"],
    ["Process", "Audit log retention",
     "How long do you keep logs?",
     "Splunk: 12 months hot + 5-year archive (S3 Glacier). PCI 12-month minimum met with margin.",
     "07 Evidence Catalog row 16"],
    ["Process", "Patch SLA",
     "CDE vs corp?",
     "CDE: 7-day Critical / 14-day High. Corp: 14-day Critical / 30-day High. Documented as separate patch policy.",
     "02 Crosswalk rows 10, 15"],
    ["Process", "DevOps gate",
     "How do you prevent PAN in dev environments?",
     "Synthetic PANs (Luhn-checked 4242 4242 4242 4242 + suffixed). DLP scan on dev env deploy. Quarterly manual review.",
     "PCI doc page 27 + TRA"],
    ["Process", "Backup encryption",
     "Backups at rest?",
     "AWS S3 + KMS-managed keys, 256-bit, annual rotation. Backups identical CDE encrypted.",
     "PCI doc page 17"],
    ["Process", "AppSec testing",
     "Static + Dynamic analysis cadence?",
     "SAST on every PR; DAST weekly; Pen-test quarterly + annual.",
     "02 Crosswalk rows 38 + 39"],
    ["Process", "Phishing defence",
     "5.4.1 v4.0 PCI anti-phishing?",
     "Proofpoint + KnowBe4 + DMARC enforcement + monthly simulation + browser-extension 'phish' detection.",
     "02 Crosswalk row 17"],
    ["Process", "Compliance dashboard",
     "What KRIs do you watch?",
     "MFA enrolment, EDR coverage, vuln SLA, awareness completion, audit log review aging, exception aging, AOC expiring ≤30d.",
     "06 Project Plan + 04 Checklist"],
]
for r, row in enumerate(QUESTIONS, 2):
    for c, v in enumerate(row, 1):
        cell = qa.cell(row=r, column=c, value=v)
        cell.alignment = cell_wrap()
        cell.border = BORDER
        if c == 2:
            cat_color = {
                "Common": EMERALD, "SOC 2": INDIGO, "PCI": AMBER, "NBFC": GOLD,
                "Process": SLATE,
            }.get(v, "")
            if cat_color:
                cell.fill = fill(cat_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [9, 14, 36, 70, 24]
for i, w in enumerate(widths, 1):
    qa.column_dimensions[get_column_letter(i)].width = w
qa.row_dimensions[1].height = 32
for r in range(2, len(QUESTIONS) + 2):
    qa.row_dimensions[r].height = 70

wb.save(OUT)
print(f"Wrote {OUT}")
