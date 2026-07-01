"""Bid scoring XLSX for SOC 2 + PCI RFP evaluation."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "Bid_Scoring_Matrix.xlsx"
wb = Workbook()

NAVY = "FF0F2A4A"; TEAL = "FF0F766E"; LIGHT = "FFF1F5F9"
THICK = Side(style="medium", color="FF0F2A4A")
THIN = Side(style="thin", color="FFCBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THIN, right=THIN, top=THICK, bottom=THICK)

def hdr_fill(): return PatternFill("solid", fgColor=NAVY)
def teal_fill(): return PatternFill("solid", fgColor=TEAL)
def light_fill(): return PatternFill("solid", fgColor=LIGHT)

def title(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = Font(size=14, bold=True, color=NAVY)

def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = Font(size=11, bold=True, color=TEAL)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill()
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

def autosize(ws):
    for col in ws.columns:
        ml = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                ml = max(ml, min(60, len(str(cell.value)) + 2))
        ws.column_dimensions[col_letter].width = max(10, ml)

# =============== Sheet 1: README ===============
ws0 = wb.active; ws0.title = "01 README"
ws0["A1"] = "Joint SOC 2 + PCI DSS Bid Scoring Matrix"
ws0["A1"].font = Font(size=14, bold=True, color=NAVY)
notes = [
    "Use this workbook to score SOC 2 and PCI DSS bid proposals from audit firms.",
    "Update the Bidder names in the Scoring & Comparison sheet (tab 4).",
    "Score each criterion 1-5 against the rubric in tabs 6-10.",
    "The Weighted Score column auto-calculates using weights in tabs 2-3.",
    "Combined score = 0.55 * SOC2 weighted + 0.45 * PCI weighted (adjustable tab 5).",
    "Final shortlist = top 3 by combined weighted score (>=70 to qualify).",
    "Lock this workbook before sending to scoring committee to avoid audit bias.",
]
for i, n in enumerate(notes, start=2):
    ws0.cell(row=i, column=1, value=n).font = Font(size=10)
ws0.column_dimensions["A"].width = 90

# =============== Sheet 2: SOC2 Weights ===============
ws1 = wb.create_sheet("02 SOC2 Weights")
section(ws1, 1, "SOC 2 Auditor weighting (must sum to 1.00)")
hdr = ["Criterion", "Weight"]
ws1.append(hdr); style_header(ws1, 1, len(hdr))
data1 = [
    ("NBFC / fintech SOC 2 track record", 0.25),
    ("Methodology & plan-of-test quality", 0.25),
    ("Key-staff quality", 0.20),
    ("Fee (absolute) vs benchmark", 0.20),
    ("Insurance + indemnity + independence", 0.10),
]
for c, w in data1:
    ws1.append([c, w])
ws1.append(["TOTAL", sum(w for _, w in data1)])
ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True)
ws1.cell(row=ws1.max_row, column=2).font = Font(bold=True)
autosize(ws1)

# =============== Sheet 3: PCI Weights ===============
ws2 = wb.create_sheet("03 PCI Weights")
section(ws2, 1, "PCI QSA weighting (must sum to 1.00)")
ws2.append(["Criterion", "Weight"]); style_header(ws2, 1, 2)
data2 = [
    ("NBFC / payment-aggregator PCI ROC track record", 0.25),
    ("v4.0.1 plan-of-test + TRA methodology", 0.25),
    ("Lead QSA + QSA-team quality", 0.20),
    ("Fee vs benchmark", 0.20),
    ("Insurance + indemnity + QSA independence", 0.10),
]
for c, w in data2:
    ws2.append([c, w])
ws2.append(["TOTAL", sum(w for _, w in data2)])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)
autosize(ws2)

# =============== Sheet 4: Scoring & Comparison ===============
ws3 = wb.create_sheet("04 Scoring Matrix")
section(ws3, 1, "Bidder scoring - update Bidder columns; weighted score auto-computed in next sheet")
hdr = ["#", "Criterion", "SOC2 weight", "PCI weight", "Bidder A - SOC2", "Bidder A - PCI",
       "Bidder B - SOC2", "Bidder B - PCI", "Bidder C - SOC2", "Bidder C - PCI"]
ws3.append(hdr); style_header(ws3, 1, len(hdr))
ws3.cell(row=2, column=5).comment = None
rows_data = [
    ("NBFC track record", 0.25, 0.25),
    ("Methodology & TRA", 0.25, 0.25),
    ("Key-staff quality", 0.20, 0.20),
    ("Fee", 0.20, 0.20),
    ("Insurance/indemnity", 0.10, 0.10),
]
for r, (c, sw, pw) in enumerate(rows_data, start=2):
    ws3.cell(row=r, column=1, value=r - 1)
    ws3.cell(row=r, column=2, value=c)
    ws3.cell(row=r, column=3, value=sw).fill = light_fill()
    ws3.cell(row=r, column=4, value=pw).fill = light_fill()
ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 35
for col_letter in ["C", "D"]:
    ws3.column_dimensions[col_letter].width = 11
for col_letter in ["E", "F", "G", "H", "I", "J"]:
    ws3.column_dimensions[col_letter].width = 14

# =============== Sheet 5: Weighted Score ===============
ws4 = wb.create_sheet("05 Weighted Score")
section(ws4, 1, "Weighted score (auto-computed)")
ws4.append(["Bidder", "SOC 2 score (0-5)", "PCI score (0-5)",
            "SOC 2 weighted", "PCI weighted", "Combined (0.55 / 0.45)"])
style_header(ws4, 1, 6)
for i, name in enumerate(["Bidder A", "Bidder B", "Bidder C"], start=2):
    ws4.cell(row=i, column=1, value=name).font = Font(bold=True)
# SOC 2 formula = SUMPRODUCT(scores, weights) / 5
ws4["B2"] = ("=SUMPRODUCT('04 Scoring Matrix'!E2:E6,'02 SOC2 Weights'!B2:B6)/5")
ws4["C2"] = ("=SUMPRODUCT('04 Scoring Matrix'!F2:F6,'03 PCI Weights'!B2:B6)/5")
ws4["D2"] = "=B2*100"
ws4["E2"] = "=C2*100"
ws4["F2"] = "=0.55*D2+0.45*E2"
ws4["B3"] = ("=SUMPRODUCT('04 Scoring Matrix'!G2:G6,'02 SOC2 Weights'!B2:B6)/5")
ws4["C3"] = ("=SUMPRODUCT('04 Scoring Matrix'!H2:H6,'03 PCI Weights'!B2:B6)/5")
ws4["D3"] = "=B3*100"
ws4["E3"] = "=C3*100"
ws4["F3"] = "=0.55*D3+0.45*E3"
ws4["B4"] = ("=SUMPRODUCT('04 Scoring Matrix'!I2:I6,'02 SOC2 Weights'!B2:B6)/5")
ws4["C4"] = ("=SUMPRODUCT('04 Scoring Matrix'!J2:J6,'03 PCI Weights'!B2:B6)/5")
ws4["D4"] = "=B4*100"
ws4["E4"] = "=C4*100"
ws4["F4"] = "=0.55*D4+0.45*E4"
for r in [2, 3, 4]:
    for c in [2, 3, 4, 5, 6]:
        ws4.cell(row=r, column=c).number_format = "0.00"
        ws4.cell(row=r, column=c).fill = light_fill()
ws4.column_dimensions["A"].width = 14
for col_letter in ["B", "C", "D", "E", "F"]:
    ws4.column_dimensions[col_letter].width = 16

# =============== Sheet 6: Rubric ===============
ws5 = wb.create_sheet("06 Rubric")
section(ws5, 1, "1-5 scoring rubric per criterion")
ws5.append(["Score", "NBFC track record", "Methodology", "Key staff",
            "Fee", "Insurance/Indemnity"])
style_header(ws5, 1, 6)
rubric = [
    ("5", ">=10 SOC2/PCI in 36 months + named NBFC refs",
     "NBFC-specific TSCs + DPDP integration + plan-of-test in 10pp",
     "Partner = ex-RBI/SEBI regulator; manager with NBFC track record",
     "<benchmark -10%", "PI 50Cr + indemnity + IC letter"),
    ("4", "6-9 SOC2/PCI in 36 months + ref list good",
     "NBFC-flavoured template", "Strong partner + manager", "<benchmark", "PI 50Cr + indemnity"),
    ("3", "3-5 SOC2/PCI in 36 months + ref list OK",
     "Generic template adapted", "Partner + manager OK", "near benchmark", "PI 25Cr"),
    ("2", "<3 SOC2/PCI in 36 months",
     "Mostly template", "Partner OK but light", ">benchmark +10%", "PI <25Cr"),
    ("1", "0 SOC2/PCI in 36 months", "No plan-of-test supplied",
     "Manager unknown", ">benchmark +20%", "No PI"),
]
for row in rubric:
    ws5.append(row)
autosize(ws5)

# =============== Sheet 7: References ===============
ws6 = wb.create_sheet("07 Reference Checks")
section(ws6, 1, "Bidder reference-call record template")
hdr = ["Bidder", "Contact name", "Client company", "Engagement year",
       "Scope", "Score (1-5)", "Red flags"]
ws6.append(hdr); style_header(ws6, 1, len(hdr))
ref_seed = [
    ("A", "<Name>", "<NBFC Client Co>", "2024", "SOC 2 + PCI", 4, ""),
    ("A", "<Name>", "<Fintech Co>", "2023", "SOC 2", 4, ""),
    ("B", "<Name>", "<NBFC Client Co>", "2024", "PCI QSA", 5, ""),
    ("C", "<Name>", "<Fintech Co>", "2023", "SOC 2", 3, "Stretched fee"),
]
for r in ref_seed:
    ws6.append(r)
autosize(ws6)

# =============== Sheet 8: Decision Log ===============
ws7 = wb.create_sheet("08 Decision Log")
section(ws7, 1, "Procurement decision log - capture all material events")
hdr = ["Date", "Action", "Owner", "Outcome"]
ws7.append(hdr); style_header(ws7, 1, len(hdr))
events = [
    ("<date>", "RFP issued (Schedule A + B + C + D)", "CFO Procurement", "Bidder portal invitations sent"),
    ("<date>", "Pre-bid Q&A session", "CISO + DPO", "FAQ published"),
    ("<date>", "Bid submission closes", "CFO Procurement", "Sealed bids logged"),
    ("<date>", "Bid presentations", "Scoring committee", "Notes distributed"),
    ("<date>", "Shortlist confirmed", "CFO + CISO", "Top 3 by combined weighted score"),
    ("<date>", "Scope walk with shortlisted bidders", "CISO + IT", "MSD finalists attached"),
    ("<date>", "Negotiation + contract close", "Procurement + Legal", "Master scope document + Schedule A signed"),
    ("<date>", "Engagement kickoff", "CFO + CISO", "Engagement partner continuity confirmed"),
]
for e in events:
    ws7.append(e)
autosize(ws7)

wb.save(OUT)
print("Wrote", OUT)
