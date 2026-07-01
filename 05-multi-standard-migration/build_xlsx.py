"""
Multi-Standard Migration Workbook
Expands the SOC 2 + PCI joint programme to include DPDP Act 2023, ISO 27001:2022, and NIST CSF 2.0.
14 sheets; ~150 rows; clean dedup logic.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

NAVY     = "1F2A44"
SLATE    = "475569"
EMERALD  = "0F766E"
AMBER    = "B45309"
RED      = "B91C1C"
INDIGO   = "3730A3"
GOLD     = "C2410C"
PAPER    = "FAFAF9"
LINE     = "D6D3D1"

THIN = Side(border_style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(color):    return PatternFill("solid", fgColor=color)
def head_font():    return Font(name="Calibri", size=11, bold=True, color="FFFFFF")
def cell_wrap():    return Alignment(wrap_text=True, vertical="top", horizontal="left")
def title_font():   return Font(name="Calibri", size=18, bold=True, color=NAVY)
def section_font(): return Font(name="Calibri", size=13, bold=True, color=NAVY)

OUT = "Multi_Standard_Programme_Workbook.xlsx"
wb = Workbook(); wb.remove(wb.active)

# ============================================================
# Sheet 1 - README
# ============================================================
cover = wb.create_sheet("01 README")
cover.column_dimensions["A"].width = 6
cover.column_dimensions["B"].width = 95
cover["A2"] = "Document"
cover["B2"] = "Multi-Standard Programme Workbook (5-Standard)"
cover["A2"].font = Font(size=10, bold=True, color=SLATE)
cover["B2"].font = title_font()

cover["A4"] = "Purpose"; cover["A4"].font = section_font()
cover["B4"] = ("Combines SOC 2 + PCI DSS v4.0 (existing joint programme) with DPDP Act 2023, "
"ISO/IEC 27001:2022, and NIST CSF 2.0. Single evidence repository, single SteerCo, single dashboard. "
"This follows the strategic decision documented in the Board Memo (NBFC/2026/CISO/JOINT-PROG/001) "
"and adds three more standards in the same programme structure.")

cover["A6"] = "Why expand"; cover["A6"].font = section_font()
cover["B6"] = ("Many NBFC clients ask for 2-3 of these standards simultaneously. A bank co-issuer asks "
"for SOC 2 + PCI + ISO 27001; an EU customer contract adds DPDP/GDPR; RBI IT Framework 2023 reads "
"almost like NIST CSF. Running five separate programmes is impossible. Running a single merged programme "
"saves 40% engineer-hours vs five separate runs.")

cover["A8"] = "Tabs in this workbook"; cover["A8"].font = section_font()
readme = [
"01 README            - this page.",
"02 Standards Map     - high-level comparison of 5 standards' purpose & audience.",
"03 Tags & Cadence    - tag legend across all 5 standards; cadence alignment.",
"04 SOC 2 <-> DPDP    - SOC 2 Privacy + DPDP Act 2023 crosswalk.",
"05 SOC 2 <-> ISO     - SOC 2 to ISO 27001:2022 crosswalk.",
"06 SOC 2 <-> NIST     - SOC 2 to NIST CSF 2.0 crosswalk.",
"07 PCI <-> ISO <-> NIST - Payment-card + ISO 27001 + NIST CSF crosswalk.",
"08 DPDP <-> ISO <-> NIST - Privacy + ISO + NIST crosswalk.",
"09 Combined Master Checklist - 100 dedup'd controls across all 5 standards.",
"10 Policy Register   - 70 documents required across the programme.",
"11 Project Plan      - 24-month multi-standard roadmap.",
"12 Evidence Catalog  - 50 artefacts cross-tagged.",
"13 Vendor Register   - vendors with SOC 2 + PCI AOC + ISO + DPDP obligations.",
"14 Multi-Audit-Trail Q&A - 40 questions covering all 5 standards.",
]
for i,r in enumerate(readme):
    cover.cell(row=9+i, column=2, value=r)

cover.column_dimensions["B"].width = 95

# ============================================================
# Sheet 2 - Standards comparison
# ============================================================
sm = wb.create_sheet("02 Standards Map")
headers = ["Standard","Geographic focus","Who cares","Cadence","Type"]
for c,h in enumerate(headers,1):
    cell = sm.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

data = [
    ["SOC 2 (AICPA SSAE-18)","Global", "Enterprise customers (banks, B2B SaaS, marketplaces)",
     "Annual Type II (12-mo window)", "Attestation report"],
    ["PCI DSS v4.0 (PCI SSC)","Global (acquirer/issuer-driven)", "Card networks, acquirer",
     "Annual ROC; quarterly ASV", "Site-data security standard + QSA attestation"],
    ["DPDP Act 2023","India", "All Indian data principals; consent+rights regime",
     "Continuous + breach 72h", "Statutory Act (data fiduciary obligations)"],
    ["ISO/IEC 27001:2022","Global", "Enterprise customers (often EU / BFSI); regulatory"
     " foundations also IMD/IRDAI", "3-year certification + annual surveillance", "Certifiable ISMS"],
    ["NIST CSF 2.0","US (but globally used)", "RBI IT Framework 2023 + US/UK B2B + auditor mappings",
     "Continuous improvement + annual maturity", "Framework profile"],
]
for r,row in enumerate(data,2):
    for c,v in enumerate(row,1):
        cell = sm.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER

widths = [12, 14, 32, 28, 22]
for i,w in enumerate(widths,1): sm.column_dimensions[get_column_letter(i)].width = w
sm.row_dimensions[1].height = 28
for r in range(2, len(data)+2): sm.row_dimensions[r].height = 60

# ============================================================
# Sheet 3 - Tags & Cadence
# ============================================================
tg = wb.create_sheet("03 Tags & Cadence")
hh = ["Tag","Meaning","NBFC example","Audit cadence","Coverage recommendation"]
for c,h in enumerate(hh,1):
    cell = tg.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

TAGS = [
    ["UNIVERSAL","Control satisfies all 5 standards.",
     "MFA everywhere, encryption, IAM, IR, audit logging",
     "Continuous+quarterly",
     "Use the strictest cadence (PCI's 12-month log; SOC 2 quarterly review)"],
    ["MOSTLY-SHARED","Covers 3-4 standards; minor gap on one.",
     "Vendor SP register (All except NIST which is framework-only)",
     "Per cycle varies",
     "Add ISO 27001 risk treatment table to satisfy"],
    ["STANDARD-SPECIFIC","Lives in one standard's domain.",
     "DPDP data principal access workflow",
     "Per standard",
     "Document separately if needed"],
    ["DIFFERENTIATED","Same activity, different cadence/evidence.",
     "Logging retention (PCI 12m vs SOC 2 flexible vs ISO unspecified)",
     "Annual + continuous",
     "Tighten to the most demanding cadence"],
]
for r,row in enumerate(TAGS,2):
    for c,v in enumerate(row,1):
        cell = tg.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==1:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [18, 32, 28, 24, 32]
for i,w in enumerate(widths,1): tg.column_dimensions[get_column_letter(i)].width = w
tg.row_dimensions[1].height = 24
for r in range(2, len(TAGS)+2): tg.row_dimensions[r].height = 80

# ============================================================
# Sheet 4 - SOC 2 <-> DPDP
# ============================================================
sd = wb.create_sheet("04 SOC 2 vs DPDP")
hh = ["SOC 2","DPDP Act 2023","Tag","Combined control","Owner","NBFC context"]
for c,h in enumerate(hh,1):
    cell = sd.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

S2_DPDP = [
    ["CC1.1 - Integrity/Ethics", "S.4 - Notice", "MOSTLY-SHARED",
     "Code of Conduct + Privacy Notice + Annual acknowledgement",
     "CISO + DPO",
     "DPDP notice is functionally distinct; combine filing."],
    ["CC2.2 - Internal Communication", "S.4 - Notice", "MOSTLY-SHARED",
     "Unified internal awareness programme with consent flow",
     "CISO + DPO", "Notice behaviour modelled per DPDP §4"],
    ["CC2.3 - External Communication", "S.6 - Purpose limitation", "STANDARD-SPECIFIC",
     "Privacy notice at layer; consent management on web/app",
     "DPO + Legal", "DPDP purposes must be SPECIFIC + lawful"],
    ["CC3.2 - Risk Assessment", "S.10 - Data fiduciary duties", "MOSTLY-SHARED",
     "Risk register with DPDP obligations for Significant Data Fiduciary",
     "CISO + DPO", "DPDP requires DPIA for high-risk processing"],
    ["CC5.3 - Policy Deployment", "S.10 - Security safeguards", "UNIVERSAL",
     "Privacy Policy + Information Security Policy merged",
     "CISO + Legal", "DPDP requires 'reasonable security'"],
    ["CC6.1 - Identity & Access", "S.10 - Security safeguards", "UNIVERSAL",
     "Same IAM+ MFA; DPDP requires 'appropriate' security",
     "CISO + IAM Lead", "Documented by TRA"],
    ["CC6.6 - Encryption (transit + rest)", "S.10 - Security safeguards", "UNIVERSAL",
     "AES-256 + TLS 1.3; documented in DPDP §10 audit",
     "Crypto Lead", "Significant Data Fiduciary encrypts"],
    ["P1.1 - Privacy notice", "S.4 - Notice", "STANDARD-SPECIFIC",
     "Consent+Receipt workflow with withdrawal endpoint",
     "DPO", "Direct DPDP mapping"],
    ["P2.1 - Purpose limitation", "S.4 + S.6 - Notice + Purpose", "STANDARD-SPECIFIC",
     "Purpose tag in data processing record + UI label",
     "DPO + Eng", "DPDP limits retention vs purpose"],
    ["P3.1 - Retention", "S.7 - Storage limitation", "STANDARD-SPECIFIC",
     "Retention rule per data class + automated purge",
     "DPO", "DPDP limits retention to purpose"],
    ["P5.1 - Data Subject Access", "S.8 - Data Principal Rights", "STANDARD-SPECIFIC",
     "Data Principal Rights workflow via portal",
     "DPO + Support", "DPDP mandates rights - access, correction, erasure, grievance"],
    ["CC7.4 - Incident Response", "DPDP S.8 - Breach notify 72h", "MOSTLY-SHARED",
     "Single IRP with DPDP 72h clock + RBI 2h + PCI acquirer 24h",
     "CISO + Legal", "DPDP requires breach notification to Board + affected data principals"],
]
for r,row in enumerate(S2_DPDP,2):
    for c,v in enumerate(row,1):
        cell = sd.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==3:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [22, 24, 16, 38, 16, 32]
for i,w in enumerate(widths,1): sd.column_dimensions[get_column_letter(i)].width = w
sd.row_dimensions[1].height = 28
for r in range(2, len(S2_DPDP)+2): sd.row_dimensions[r].height = 50

# ============================================================
# Sheet 5 - SOC 2 <-> ISO 27001
# ============================================================
si = wb.create_sheet("05 SOC 2 vs ISO 27001")
hh = ["SOC 2","ISO 27001 Annex A control","Tag","Combined control","NBFC note"]
for c,h in enumerate(hh,1):
    cell = si.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

S2_ISO = [
    ["CC1.1 - Integrity/Ethics", "A.5.1 Policies for information security", "UNIVERSAL",
     "Single Information Security Policy", "ISO 27001 anchors policy hierarchy"],
    ["CC1.2 - Board Oversight", "A.5.2 Information security roles", "MOSTLY-SHARED",
     "Single cyber committee", "ISO requires role definitions"],
    ["CC2.x - Communication", "A.5.11 Acceptable use of information assets", "MOSTLY-SHARED",
     "Policy library + acknowledgement", "ISO requires documented communication"],
    ["CC3.2 - Risk Assessment", "A.6.1.2 Information security risk assessment", "UNIVERSAL",
     "Unified risk methodology; ISO mandatory ART + CRR", "ISO needs Asset Register + Risk Treatment"],
    ["CC4.x - Monitoring", "A.5.35 Independent review", "MOSTLY-SHARED",
     "Internal audit + CCO dashboard", "ISO needs independent review"],
    ["CC5.2 - Tech controls", "A.8.9 Configuration management", "UNIVERSAL",
     "Same baseline config procedure + change management", "ISO Annex A now has tech controls"],
    ["CC6.1 - Logical Access", "A.8.2 Privileged access rights + A.8.5 Secure authentication", "UNIVERSAL",
     "Same SSO + MFA; ISO adds PAM for privileged access", "ISO mandatory criteria on MFA"],
    ["CC6.2 - New user access", "A.8.2(a),(b)", "UNIVERSAL",
     "JML workflow",
     "ISO requires provision + de-provision process"],
    ["CC6.6 - Encryption", "A.8.24 Cryptography", "UNIVERSAL",
     "Same KMS + HSM; ISO requires CRY key management documented", "ISO prescribes use of crypto"],
    ["CC7.1 - Vulnerability", "A.8.8 Management of technical vulnerabilities", "UNIVERSAL",
     "Same scanner + SLA",
     "ISO requires vuln schedule"],
    ["CC7.4 - Incident Response", "A.5.24 Information security incident management planning",
     "UNIVERSAL", "Same IR plan + tabletops",
     "ISO mandatory document & plan"],
    ["CC7.5 - BCP / DR", "A.5.30 ICT readiness for business continuity", "UNIVERSAL",
     "Same DR plan + tests", "ISO mandatory"],
    ["CC8.1 - Change management", "A.8.32 Change management", "UNIVERSAL",
     "Same CAB + records",
     "ISO requires change procedure"],
    ["CC9.2 - Vendor risk", "A.5.19 Information security in supplier relationships",
     "MOSTLY-SHARED", "Single SP register + AOC",
     "ISO requires supplier agreements + assessment"],
    ["A1.3 - DR Test", "A.5.30 + A.8.14 Testing", "UNIVERSAL",
     "Annual DR drill evidence",
     "ISO mandatory; SOC 2 also"],
    ["C1.x - Confidentiality", "A.5.12 Classification of information + A.8.10 Information deletion",
     "UNIVERSAL", "Same data classification scheme + tagging",
     "NBFC: PAN = Restricted top tier"],
    ["PI1.x - Processing integrity", "A.8.28 Secure coding", "MOSTLY-SHARED",
     "Secure SDLC + change control",
     "ISO requires secure coding"],
    ["P1.x - Privacy", "A.5.34 Privacy & PII protection", "STANDARD-SPECIFIC",
     "Privacy notice + consent workflow (separate from DPDP)",
     "ISO vs DPDP both require notice"],
]
for r,row in enumerate(S2_ISO,2):
    for c,v in enumerate(row,1):
        cell = si.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==3:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [22, 30, 14, 38, 30]
for i,w in enumerate(widths,1): si.column_dimensions[get_column_letter(i)].width = w
si.row_dimensions[1].height = 28
for r in range(2, len(S2_ISO)+2): si.row_dimensions[r].height = 50

# ============================================================
# Sheet 6 - SOC 2 <-> NIST CSF 2.0
# ============================================================
sn = wb.create_sheet("06 SOC 2 vs NIST CSF 2.0")
hh = ["SOC 2","NIST CSF Function / Category","Tag","Combined control","NBFC note"]
for c,h in enumerate(hh,1):
    cell = sn.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

S2_NIST = [
    ["CC1.x - Governance","GV (Govern): organisational context and risk strategy","UNIVERSAL",
     "Single cyber committee + governance charter","NIST 2.0 added GV - perfect SOC 2 CC1.x overlap"],
    ["CC2.x - Comms","GV.OC (governance in org context); PR.AT (awareness training)","UNIVERSAL",
     "Same internal + external comms + training","NIST PR.AT mirrors CC2"],
    ["CC3.2 - Risk Assessment","GV.RM (risk management strategy); ID.RA (risk assessment)","UNIVERSAL",
     "Same risk register","NIST 2.0 aligned with ISO 27005"],
    ["CC4.1-2 - Monitoring","DE.CM (continuous monitoring); ID.AM (asset management)","UNIVERSAL",
     "Same KPI dashboard","NIST provides inventory category"],
    ["CC5.3 - Policies","GV.PO (organisational policies)","UNIVERSAL",
     "Same policy library","NIST 2.0 formalised policies"],
    ["CC6.1 - Logical Access","PR.AA (Identity, Authentication, Access Control)","UNIVERSAL",
     "Same IAM + MFA + RBAC","NIST 2.0 redesign of PR.AC"],
    ["CC6.6 - Encryption","PR.DS (Data Security)","UNIVERSAL",
     "Same crypto + HSM","NIST 2.0 expanded DS"],
    ["CC6.8 - Malware","PR.PS (Platform security) - malware","UNIVERSAL",
     "Same EDR","NIST splits platform security"],
    ["CC7.1 - Vulnerability","ID.RA + DE.CM - vulnerabilities","UNIVERSAL",
     "Same vuln programme","NIST provides explicit category"],
    ["CC7.2 - Monitoring","DE.AE (anomalies and events); DE.CM","UNIVERSAL",
     "Same SIEM","NIST 2.0 splits DE"],
    ["CC7.4 - Incident Response","RS.RP (response planning); RS.MI (mitigation); RS.CO (comm)","UNIVERSAL",
     "Same IRP + runbooks + tabletop","NIST 2.0 more structured than v1.1"],
    ["CC7.5 / A1.3 - Recover","RC.RP (recovery planning); RC.IM (improvement)","UNIVERSAL",
     "Same DR plan + lessons-learned","NIST's RC matches SOC 2 + ISO"],
    ["CC9.2 - Vendor risk","GV.SC (cyber supply chain risk management)","UNIVERSAL",
     "Same SP register + AOC","NIST 2.0 added GV.SC"],
    ["CC9.1 - Insurance - not NIST","n/a","MOSTLY-SHARED",
     "Cyber insurance policy + RBI requirement",
     "SOC 2 CC9.1; no NIST equivalent but document separately"],
]
for r,row in enumerate(S2_NIST,2):
    for c,v in enumerate(row,1):
        cell = sn.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==3:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [22, 30, 14, 38, 32]
for i,w in enumerate(widths,1): sn.column_dimensions[get_column_letter(i)].width = w
sn.row_dimensions[1].height = 28
for r in range(2, len(S2_NIST)+2): sn.row_dimensions[r].height = 50

# ============================================================
# Sheet 7 - PCI <-> ISO <-> NIST
# ============================================================
pn = wb.create_sheet("07 PCI vs ISO vs NIST")
hh = ["PCI Req","ISO 27001","NIST CSF 2.0","Tag","Combined","NBFC note"]
for c,h in enumerate(hh,1):
    cell = pn.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER
PCI_ISO_NIST = [
    ["Req 1 - Network Security Controls","A.8.20 Network security; A.8.22 Segregation",
     "PR.IR (Infrastructure resilience); PR.PS - network","MOSTLY-SHARED",
     "Single network policy + segmentation","NBFC: CDE VPC isolated"],
    ["Req 3 - Protect stored data","A.8.24 Cryptography + A.8.10 Deletion",
     "PR.DS (Data Security)","MOSTLY-SHARED",
     "Card vault encryption + KMS + HSM","PCI specifics for PAN handling"],
    ["Req 4 - Transmission","A.8.24 Cryptography (transport)",
     "PR.DS","MOSTLY-SHARED",
     "TLS 1.2+ enforced","Same TLS standard ACL across NBFC"],
    ["Req 5 - Anti-malware","A.8.7 Protection against malware",
     "PR.PS","UNIVERSAL",
     "Same EDR + Linux","CrowdStrike / SentinelOne everywhere"],
    ["Req 6 - Secure systems","A.8.28 Secure coding + A.8.9 Configuration management",
     "PR.PS + PR.DV - secure dev","MOSTLY-SHARED",
     "Same SDLC + code review","Testing on payment flows"],
    ["Req 7 - Restrict by need","A.5.15 Access control + A.8.2",
     "PR.AA","UNIVERSAL",
     "Same RBAC","Card vault access strictly limited"],
    ["Req 8 - Auth + MFA","A.5.17 Authentication info + A.5.18 Access rights",
     "PR.AA - MFA","MOSTLY-SHARED",
     "Same SSO + WebAuthn + 3DS2","WebAuthn staff / 3DS2 customers"],
    ["Req 9 - Physical","A.7 Physical security",
     "PR.AA - physical","MOSTLY-SHARED",
     "Same physical security procedure","PCI adds POI specifics"],
    ["Req 10 - Logging","A.8.15 Logging + A.5.28 Evidence collection",
     "DE.AE + DE.CM","UNIVERSAL",
     "Same SIEM + log retention","PCI 12-mo; ISO unspecified; NIST unspecified - safer to use PCI"],
    ["Req 11 - Test regularly","A.8.8 Vuln mgmt + A.8.16 Monitoring + A.5.35 Review",
     "ID.RA + DE.CM","UNIVERSAL",
     "Same pen-test + scans + ASV","Single vendor for both"],
    ["Req 12 - Org policies","A.5.1 to A.5.39 - organisational policies (entire A.5)",
     "GV.PO","UNIVERSAL",
     "Same policy library","Nearly identical coverage"],
]
for r,row in enumerate(PCI_ISO_NIST,2):
    for c,v in enumerate(row,1):
        cell = pn.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==4:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [26, 30, 24, 14, 36, 30]
for i,w in enumerate(widths,1): pn.column_dimensions[get_column_letter(i)].width = w
pn.row_dimensions[1].height = 28
for r in range(2, len(PCI_ISO_NIST)+2): pn.row_dimensions[r].height = 60

# ============================================================
# Sheet 8 - DPDP <-> ISO <-> NIST
# ============================================================
dn = wb.create_sheet("08 DPDP vs ISO vs NIST")
hh = ["DPDP Act §","ISO 27001","NIST CSF 2.0","Tag","Combined","NBFC note"]
for c,h in enumerate(hh,1):
    cell = dn.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER
DPDP_ISO_NIST = [
    ["§4 Notice","A.5.34 Privacy + A.5.32 IP rights","GV.OC (organisational context)","STANDARD-SPECIFIC",
     "Privacy notice + consent receipt","DPDP sharper; ISO generic"],
    ["§6 Purpose limitation","A.5.34","ID.AM asset mgmt","STANDARD-SPECIFIC",
     "Purpose tag in data classification","DPDP requires SPECIFIC purposes"],
    ["§7 Storage limitation","A.8.10 Information deletion","PR.IP","STANDARD-SPECIFIC",
     "Auto-purge + retention register","Note DPDP max retention"],
    ["§8 Data Principal Rights","A.5.34","GV.OC","STANDARD-SPECIFIC",
     "Portal + ticketing workflow","DPDP legally enforceable rights"],
    ["§10 Security safeguards","A.5.15-A.5.18 + A.8.24","PR.AA + PR.DS","UNIVERSAL",
     "Same IAM + crypto + SIEM","CISO + DPO alignment"],
    ["§13 DPIA","A.5.7 Threat intel + A.5.30","GV.RM + ID.RA","STANDARD-SPECIFIC",
     "DPIA + TRA shared template","Significant Data Fiduciary required"],
    ["§14 Cross-border","A.5.21 Managing info security in supplier relationships","GV.SC","STANDARD-SPECIFIC",
     "Region-lock + SCP rules","RBI prohibits to non-notified countries"],
    ["§72h breach","A.5.24-A.5.27 Incident mgmt","RS.CO","UNIVERSAL",
     "Single IR plan with DPDP 72h clock","Multiple clock triggers; one IRP"],
    ["§16 Cross-border transfer","A.8.21 Data leakage prevention","PR.IP","STANDARD-SPECIFIC",
     "DSPM + region-lock","Region enforcement policy"],
    ["§26 Penalties","Internal compliance","GV.RM (verification)","MOSTLY-SHARED",
     "Compliance dashboard + internal audit","₹250 crore penalty avoidance"],
]
for r,row in enumerate(DPDP_ISO_NIST,2):
    for c,v in enumerate(row,1):
        cell = dn.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==4:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [22, 30, 24, 14, 36, 30]
for i,w in enumerate(widths,1): dn.column_dimensions[get_column_letter(i)].width = w
dn.row_dimensions[1].height = 28
for r in range(2, len(DPDP_ISO_NIST)+2): dn.row_dimensions[r].height = 60

# ============================================================
# Sheet 9 - Combined Master Checklist (100 rows)
# ============================================================
mc = wb.create_sheet("09 Master Checklist")
mc_h = ["#","Domain","Control","Tag","Standards covered","Cadence","Owner"]
for c,h in enumerate(mc_h,1):
    cell = mc.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

CHECKLIST = [
    # Governance
    [1,"Governance","Information Security Policy","UNIVERSAL","All 5","Annual","CISO"],
    [2,"Governance","Cyber Committee charter + Board minutes","UNIVERSAL","SOC 2 + ISO + NIST (RBI)","Quarterly","CISO"],
    [3,"Governance","Risk assessment methodology","UNIVERSAL","SOC 2 + PCI + ISO + NIST + DPDP","Annual","CRO + CISO"],
    [4,"Governance","Cybersecurity operational risk register","UNIVERSAL","SOC 2 + ISO + NIST + DPDP","Quarterly","CISO"],
    [5,"Governance","Statement of Applicability (SoA)","STANDARD-SPECIFIC","ISO 27001 only","Annual","CISO"],
    [6,"Governance","Risk Treatment Plan (RTP)","STANDARD-SPECIFIC","ISO 27001 only","Annual","CISO"],
    [7,"Governance","Internal audit programme + charter","UNIVERSAL","SOC 2 + ISO + NIST","Annual","Internal Audit"],
    [8,"Governance","Management review (MRM) for ISMS","STANDARD-SPECIFIC","ISO 27001 only","Quarterly","CISO"],
    # Access
    [9,"IAM","Identity provider (Okta/Entra) with MFA","UNIVERSAL","All 5","Continuous","IAM Lead"],
    [10,"IAM","Privileged Access Management","UNIVERSAL","SOC 2 + ISO + PCI + NIST","Continuous","IAM Lead"],
    [11,"IAM","FIDO2 / WebAuthn phishing-resistant MFA","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","IAM Lead"],
    [12,"IAM","Joiners / Movers / Leavers workflow","UNIVERSAL","All 5","Continuous","IT + HR"],
    [13,"IAM","Quarterly access reviews","MOSTLY-SHARED","SOC 2 + ISO + PCI + NIST","Quarterly","Business unit heads"],
    [14,"IAM","RBAC + Just-In-Time elevation","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","IAM Lead"],
    [15,"IAM","Service account inventory + Secrets Manager","UNIVERSAL","SOC 2 + PCI + IAM + ISO + NIST","Continuous","Cloud + AppSec"],
    # Cryptography
    [16,"Crypto","Encryption at rest (AES-256, KMS managed)","UNIVERSAL","All 5","Continuous","Crypto Lead"],
    [17,"Crypto","TLS 1.2+ enforced at perimeter","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","Network Lead"],
    [18,"Crypto","Annual key rotation + dual control","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Annual","Crypto Lead"],
    [19,"Crypto","HSM (CloudHSM FIPS 140-2 L3)","MOSTLY-SHARED","PCI + ISO + NIST","Continuous","Crypto Lead"],
    [20,"Crypto","Card vault tokenisation (VTS+MDES+R-T-V)","STANDARD-SPECIFIC","RBI Tokenisation","Continuous","CTO"],
    # Network
    [21,"Network","Network Security Controls (NSC)","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","Network Lead"],
    [22,"Network","Default-deny firewall rules","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","Network Lead"],
    [23,"Network","Segmentation of CDE","MOSTLY-SHARED","PCI + ISO + NIST","Continuous","Network Lead"],
    [24,"Network","WAF at perimeter + payment route","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Continuous","AppSec"],
    [25,"Network","IDS / IPS at perimeter","MOSTLY-SHARED","SOC 2 + ISO + NIST","Continuous","SecOps"],
    [26,"Network","TLS scan quarterly","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Quarterly","SecOps"],
    # AppSec
    [27,"AppSec","Secure SDLC policy","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Annual","CTO"],
    [28,"AppSec","SAST / SCA per build","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Continuous","AppSec"],
    [29,"AppSec","DAST weekly","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Weekly","AppSec"],
    [30,"AppSec","3DS2 for online card transactions","STANDARD-SPECIFIC","PCI + RBI tokenisation","Continuous","Payments Eng"],
    [31,"AppSec","Script inventory + Subresource Integrity","STANDARD-SPECIFIC","PCI 6.4.3 / 11.6.1","Weekly","AppSec"],
    [32,"AppSec","Ransomware simulation","STANDARD-SPECIFIC","RBI 2h breach","Annual","CISO"],
    [33,"AppSec","Penetration test (annual + on-change)","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Annual","AppSec"],
    [34,"AppSec","Software Bill of Materials (SBOM)","STANDARD-SPECIFIC","PCI 6.3.3","Monthly","AppSec"],
    # SysOps / Monitoring
    [35,"SysOps","Centralised logging (SIEM)","UNIVERSAL","SOC 2 + PCI + NIST","Continuous","SOC Lead"],
    [36,"SysOps","12-month online log retention","DIFFERENTIATED","PCI strictest","Continuous","SOC Lead"],
    [37,"SysOps","File Integrity Monitoring (FIM)","MOSTLY-SHARED","SOC 2 + PCI + NIST","Weekly","SecOps"],
    [38,"SysOps","EDR everywhere incl. Linux","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","IT"],
    [39,"SysOps","Email security + anti-phishing","MOSTLY-SHARED","SOC 2 + PCI 5.4.1 + ISO + NIST","Continuous","CISO + IT"],
    [40,"SysOps","Backup with AES-256 encryption + quarterly restore test","UNIVERSAL","All 5","Quarterly","IT"],
    [41,"SysOps","Vulnerability scan (monthly)","UNIVERSAL","SOC 2 + ISO + NIST + PCI","Monthly","SecOps"],
    [42,"SysOps","Authenticated internal vulnerability scan","MOSTLY-SHARED","PCI 11.3.1.2 + SOC 2","Quarterly","SecOps"],
    [43,"SysOps","Quarterly ASV scan (external)","STANDARD-SPECIFIC","PCI only","Quarterly","SecOps"],
    [44,"SysOps","Daily reconciliation of disbursements","STANDARD-SPECIFIC","SOC 2 PI + PCI Req 10","Daily","Finance + Eng"],
    # IR & DR
    [45,"IR","Documented Incident Response Plan","UNIVERSAL","SOC 2 + PCI + NIST + RBI","Annual","CISO"],
    [46,"IR","Incident Response runbooks (P1-P4)","UNIVERSAL","SOC 2 + PCI + RBI","Per incident","SOC"],
    [47,"IR","RBI 2-hour notification mandate","STANDARD-SPECIFIC","RBI","Per incident","CISO + Legal"],
    [48,"IR","PCI acquirer notification within 24h","STANDARD-SPECIFIC","PCI","Per incident","CISO + Legal"],
    [49,"IR","DPDP 72h breach notification","STANDARD-SPECIFIC","DPDP","Per incident","DPO + Legal"],
    [50,"IR","CERT-In 6h notification","STANDARD-SPECIFIC","CERT-In","Per incident","CISO"],
    [51,"IR","Tabletop exercise 2x/year","UNIVERSAL","SOC 2 + ISO + NIST + RBI","Bi-annual","CISO"],
    [52,"DR","BCP / DR plan","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Annual","CIO + CISO"],
    [53,"DR","Annual DR drill + RTO/RPO captured","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Annual","IT + CISO"],
    [54,"DR","Backup integrity verification","MOSTLY-SHARED","SOC 2 + PCI + ISO","Quarterly","IT"],
    # Change Mgmt
    [55,"Change","Change Management Policy + CAB","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","Engineering"],
    [56,"Change","Production access via JIT elevation","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","IAM + Engineering"],
    [57,"Change","Emergency change SOP","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Continuous","Engineering"],
    [58,"Change","Pre-prod → Prod segregation","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","Engineering"],
    # Vendor
    [59,"Vendor","Service Provider Register","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Quarterly","Compliance + CISO"],
    [60,"Vendor","Vendor SOC 2 + PCI AOC + ISO 27001 certs","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Annual","Compliance"],
    [61,"Vendor","MSA with PCI acknowledgement clause","STANDARD-SPECIFIC","PCI 12.8","Per vendor","Legal"],
    [62,"Vendor","Data Processing Agreement (DPA)","STANDARD-SPECIFIC","DPDP §10 + EU GDPR","Per vendor","Legal + DPO"],
    [63,"Vendor","Annual SOC 2 + PCI AOC expiry alerts","UNIVERSAL","SOC 2 + PCI + ISO","Continuous","Compliance"],
    # Privacy
    [64,"Privacy","Privacy notice (web + mobile)","STANDARD-SPECIFIC","DPDP §4 + ISO A.5.34 + SOC 2 P1","Annual","DPO + Legal"],
    [65,"Privacy","Consent receipt per data principal","STANDARD-SPECIFIC","DPDP §6","Continuous","DPO + Eng"],
    [66,"Privacy","Data Principal Rights workflow","STANDARD-SPECIFIC","DPDP §8","Continuous","DPO + Support"],
    [67,"Privacy","Records of Processing Activities (RoPA)","STANDARD-SPECIFIC","DPDP §11 + ISO A.5.34","Quarterly","DPO"],
    [68,"Privacy","Cross-border transfer control (region lock)","STANDARD-SPECIFIC","DPDP §14","Continuous","Cloud Lead"],
    # Audit
    [69,"Audit","Internal audit plan with 5-standard mapping","UNIVERSAL","SOC 2 + PCI + ISO + NIST + DPDP","Annual","Internal Audit"],
    [70,"Audit","Quarterly board pack","MOSTLY-SHARED","SOC 2 + PCI + ISO + RBI","Quarterly","CISO + CRO"],
    # RBI-specific
    [71,"RBI","RBI Master Direction on Outsourcing 2021","STANDARD-SPECIFIC","RBI","Annual","Compliance"],
    [72,"RBI","RBI IT Framework for NBFC 2023","STANDARD-SPECIFIC","RBI","Annual","CISO"],
    [73,"RBI","RBI Cyber Security Framework 2023","STANDARD-SPECIFIC","RBI","Continuous","CISO"],
    [74,"RBI","RBI Tokenisation Guidelines 2021","STANDARD-SPECIFIC","RBI tokenisation","Continuous","CTO"],
    [75,"RBI","RBI Digital Lending Guidelines 2022","STANDARD-SPECIFIC","RBI digital lending","Continuous","Product"],
    # Awareness/Training
    [76,"Training","Annual security training","UNIVERSAL","SOC 2 + ISO + NIST + RBI","Annual","CISO + HR"],
    [77,"Training","Quarterly phishing simulation","MOSTLY-SHARED","SOC 2 + PCI 5.4.1 + ISO + NIST","Quarterly","CISO + HR"],
    # Physical
    [78,"Physical","Badge + biometrics at facilities","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","Facilities + CISO"],
    [79,"Physical","Visitor management","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","Facilities"],
    [80,"Physical","POI / POS device monthly inspection","STANDARD-SPECIFIC","PCI Req 9","Monthly","Operations"],
    # Data Lifecycle
    [81,"Data Lifecycle","Data classification scheme","UNIVERSAL","SOC 2 + ISO + NIST + DPDP","Quarterly","DPO + CISO"],
    [82,"Data Lifecycle","Cardholder data storage inventory","STANDARD-SPECIFIC","PCI Req 3","Quarterly","DPO"],
    [83,"Data Lifecycle","Auto-purge per retention","MOSTLY-SHARED","DPDP §7 + ISO A.8.10","Continuous","Eng + DPO"],
    # AI / LLM
    [84,"AI","AI/LLM risk assessment","MOSTLY-SHARED","DPDP + NIST + emerging EU AI Act","Annual","Product Security"],
    [85,"AI","Prompt injection defence controls","STANDARD-SPECIFIC","AI Act + OWASP Top 10 LLM","Quarterly","Product Security"],
    # Compliance programme
    [86,"Prog","SteerCo charter","STANDARD-SPECIFIC","All standards governance","Annual","CISO + CRO"],
    [87,"Prog","GRC tooling (Vanta / Drata / Tugboat)","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Continuous","CISO + IT"],
    [88,"Prog","Joint dashboard","UNIVERSAL","All 5","Continuous","Compliance"],
    # OT/ICS
    [89,"OT","OT/ICS risk register","MOSTLY-SHARED","ISO + NIST","Annual","CISO"],
    # Quantum
    [90,"Quantum","Crypto agility roadmap (post-quantum)","STANDARD-SPECIFIC","NIST + emerging","Annual","Crypto Lead"],
    # Cloud
    [91,"Cloud","Cloud Security Posture Management (CSPM)","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Continuous","Cloud Lead"],
    [92,"Cloud","AWS Config rules active","MOSTLY-SHARED","SOC 2 + PCI + ISO + NIST","Continuous","Cloud Lead"],
    # Plan
    [93,"Plan","Tabletop every 6 months","UNIVERSAL","SOC 2 + ISO + NIST + PCI","Bi-annual","SOC"],
    [94,"Plan","DR drill annually with documented RTO/RPO","UNIVERSAL","SOC 2 + PCI + ISO + NIST","Annual","IT + CISO"],
    [95,"Plan","Cybersecurity Committee meets Q","MOSTLY-SHARED","RBI §13 + ISO","Quarterly","Board"],
    # Compliance audit
    [96,"Audit","SOC 2 Type II annual","STANDARD-SPECIFIC","SOC 2 only","Annual","CFO + CISO"],
    [97,"Audit","PCI ROC annual","STANDARD-SPECIFIC","PCI only","Annual","CFO + CISO + QSA"],
    [98,"Audit","ISO 27001 surveillance annual","STANDARD-SPECIFIC","ISO only","Annual","CFO + CISO + CB"],
    [99,"Audit","NIST CSF maturity assessment annual","STANDARD-SPECIFIC","NIST only","Annual","CISO + Internal Audit"],
    [100,"Audit","DPDP compliance self-attestation","STANDARD-SPECIFIC","DPDP","Annual","DPO + CISO"],
]
for r,row in enumerate(CHECKLIST,2):
    for c,v in enumerate(row,1):
        cell = mc.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==4:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [4, 18, 50, 14, 30, 14, 22]
for i,w in enumerate(widths,1): mc.column_dimensions[get_column_letter(i)].width = w
mc.row_dimensions[1].height = 28
for r in range(2, len(CHECKLIST)+2): mc.row_dimensions[r].height = 35
mc.freeze_panes = "A2"

# ============================================================
# Sheet 10 - Policy Register (70 docs)
# ============================================================
pr = wb.create_sheet("10 Policy Register")
hh = ["#","Document","Standards covered","Tag","Owner","Cadence"]
for c,h in enumerate(hh,1):
    cell = pr.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

POLICIES = [
    [1,"Information Security Policy","All 5","UNIVERSAL","CISO","Annual"],
    [2,"Risk Management Policy","All 5","UNIVERSAL","CRO","Annual"],
    [3,"Acceptable Use Policy","SOC 2 + ISO + NIST","UNIVERSAL","CISO","Annual"],
    [4,"Code of Conduct","UNIVERSAL","UNIVERSAL","HR","Annual"],
    [5,"Background Verification Policy","SOC 2 + PCI + DPDP","UNIVERSAL","HR","On-hire"],
    [6,"Access Control Policy","All 5","UNIVERSAL","CISO","Annual"],
    [7,"Authentication / MFA Policy","All 5","UNIVERSAL","CISO + IAM","Annual"],
    [8,"Cryptographic Standards","All 5","UNIVERSAL","Crypto Lead","Annual + on change"],
    [9,"Key Management Policy","SOC 2 + PCI + ISO","MOSTLY-SHARED","Crypto Lead","Annual"],
    [10,"Data Classification Policy","All 5","UNIVERSAL","DPO + CISO","Annual"],
    [11,"Privacy Notice & Policy","DPDP + SOC 2 + ISO","STANDARD-SPECIFIC","DPO + Legal","Annual"],
    [12,"Consent Receipt Procedure","DPDP only","STANDARD-SPECIFIC","DPO","Continuous"],
    [13,"Data Principal Rights Workflow","DPDP only","STANDARD-SPECIFIC","DPO + Support","Continuous"],
    [14,"Cross-border Transfer Procedure","DPDP only","STANDARD-SPECIFIC","DPO + Cloud","Continuous"],
    [15,"Card Data Storage + Retention","PCI only","STANDARD-SPECIFIC","CISO + Compliance","Annual"],
    [16,"Cardholder Data Protection Policy","PCI only","STANDARD-SPECIFIC","CISO","Annual"],
    [17,"Network Security Policy","All 5","UNIVERSAL","Network Lead + CISO","Annual"],
    [18,"Firewall / NSC Review Procedure","SOC 2 + PCI + ISO","UNIVERSAL","Network Lead","Bi-annual"],
    [19,"Secure Configuration Standards","SOC 2 + PCI + ISO + NIST","UNIVERSAL","Infra Lead","Annual"],
    [20,"Patch Management Policy","All 5","UNIVERSAL","Infra Lead","Continuous"],
    [21,"Anti-Malware (EDR) Policy","All 5","UNIVERSAL","CISO","Annual"],
    [22,"Anti-Phishing Programme","SOC 2 + PCI 5.4.1 + NIST","MOSTLY-SHARED","CISO","Quarterly"],
    [23,"Secure SDLC Policy","SOC 2 + PCI + ISO + NIST","UNIVERSAL","CTO","Annual + on change"],
    [24,"Payment Page Script Inventory Procedure","PCI 6.4.3 only","STANDARD-SPECIFIC","AppSec","Quarterly"],
    [25,"Change Management Policy","All 5","UNIVERSAL","Engineering","Annual"],
    [26,"Logging & Monitoring Policy","SOC 2 + PCI + ISO + NIST","UNIVERSAL","CISO + SOC","Annual"],
    [27,"Audit Log Review Procedure","PCI + SOC 2","MOSTLY-SHARED","SOC Manager","Daily + Weekly"],
    [28,"Time Synchronisation Procedure","SOC 2 + PCI + ISO","UNIVERSAL","Infra","Annual"],
    [29,"Vulnerability Management Policy","All 5","UNIVERSAL","CISO","Annual"],
    [30,"ASV Scan Procedure","PCI only","STANDARD-SPECIFIC","SecOps","Quarterly"],
    [31,"Internal Vulnerability Scan Procedure","SOC 2 + PCI + ISO + NIST","UNIVERSAL","SecOps","Quarterly"],
    [32,"File Integrity Monitoring Procedure","SOC 2 + PCI + NIST","MOSTLY-SHARED","SecOps","Weekly"],
    [33,"Penetration Test Charter","SOC 2 + PCI + ISO + NIST","UNIVERSAL","AppSec","Annual"],
    [34,"Wireless + Rogue Device Procedure","PCI 11.2 only","STANDARD-SPECIFIC","SecOps","Annual"],
    [35,"Physical Security Policy (DC)","All 5","UNIVERSAL","Facilities + CISO","Annual"],
    [36,"Visitor Management Procedure","All 5","UNIVERSAL","Facilities","Daily"],
    [37,"POI / POS Device Inventory Procedure","PCI 9.5 only","STANDARD-SPECIFIC","Operations","Monthly"],
    [38,"Media Handling & Destruction","SOC 2 + PCI + ISO","UNIVERSAL","IT + Facilities","Continuous"],
    [39,"Security Awareness Training Policy","All 5","UNIVERSAL","CISO + HR","Annual"],
    [40,"Mobile App Security Policy","SOC 2 + PCI + DPDP","UNIVERSAL","Mobile Security","Annual"],
    [41,"Personnel Security Policy","SOC 2 + PCI + ISO","UNIVERSAL","HR + CISO","Annual"],
    [42,"Service Provider Management Policy","SOC 2 + PCI + ISO","UNIVERSAL","CISO + Procurement","Annual"],
    [43,"Service Provider Register","All 5","UNIVERSAL","Compliance","Quarterly"],
    [44,"MSA with PCI Acknowledgement","PCI only","STANDARD-SPECIFIC","Legal + Compliance","Per vendor"],
    [45,"Data Processing Agreement (DPA)","DPDP + ISO + SOC 2","MOSTLY-SHARED","Legal + DPO","Per vendor"],
    [46,"Incident Response Plan","All 5","UNIVERSAL","CISO","Annual"],
    [47,"IR Runbooks (P1-P4)","SOC 2 + PCI + RBI","MOSTLY-SHARED","SOC Manager","Per incident"],
    [48,"RBI 2-hour Notification Procedure","RBI only","STANDARD-SPECIFIC","CISO + Legal","Per incident"],
    [49,"DPDP 72-hour Breach Notification","DPDP only","STANDARD-SPECIFIC","DPO + Legal","Per incident"],
    [50,"BCP / DR Plan","All 5","UNIVERSAL","CIO + CISO","Annual"],
    [51,"Risk Assessment Methodology","All 5","UNIVERSAL","CRO + CISO","Annual"],
    [52,"Targeted Risk Analysis (TRA) Templates","PCI 4.0 only","STANDARD-SPECIFIC","CISO","Annual"],
    [53,"PCI Compliance Programme Summary","PCI only","STANDARD-SPECIFIC","Compliance","Annual"],
    [54,"SOC 2 Type II Programme Summary","SOC 2 only","STANDARD-SPECIFIC","Compliance","Annual"],
    [55,"ISO 27001 SoA","ISO only","STANDARD-SPECIFIC","CISO","Annual"],
    [56,"ISO 27001 RTP","ISO only","STANDARD-SPECIFIC","CISO","Annual"],
    [57,"Statement of Applicability","ISO only","STANDARD-SPECIFIC","CISO","Annual + on change"],
    [58,"Annual Compliance Maintenance Schedule","All 5","UNIVERSAL","Compliance","Annual"],
    [59,"Quarterly Joint Compliance Dashboard","All 5","UNIVERSAL","Compliance","Quarterly"],
    [60,"Cybersecurity Committee Charter","SOC 2 + RBI","MOSTLY-SHARED","Board Secretariat","Annual"],
    [61,"BYOD Policy","SOC 2 + ISO + NIST","UNIVERSAL","HR + CISO","Annual"],
    [62,"Work-from-Anywhere Policy","SOC 2 + ISO + NIST","UNIVERSAL","HR + IT + CISO","Annual"],
    [63,"Open Source Usage Policy","SOC 2 + ISO","UNIVERSAL","CTO + Legal","Annual"],
    [64,"Cyber Insurance Policy + Renewal","SOC 2 only","STANDARD-SPECIFIC","CFO + Legal","Annual"],
    [65,"AI/LLM Use Risk Policy","ISO + DPDP + NIST","MOSTLY-SHARED","Product Security","Annual"],
    [66,"OT/ICS Security Policy","ISO + NIST","MOSTLY-SHARED","CISO","Annual"],
    [67,"Post-quantum Crypto Migration","NIST","STANDARD-SPECIFIC","Crypto Lead","Annual"],
    [68,"Cloud Security Policy (CSPM)","All 5","UNIVERSAL","Cloud Lead + CISO","Annual"],
    [69,"Outsourcing Policy (NBFC RBI)","RBI","STANDARD-SPECIFIC","Compliance + Legal","Annual"],
    [70,"26AB Grievance Redressal (NBFC)","RBI","STANDARD-SPECIFIC","Customer Success","Annual"],
]
for r,row in enumerate(POLICIES,2):
    for c,v in enumerate(row,1):
        cell = pr.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==4:
            tag_color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if tag_color:
                cell.fill = fill(tag_color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [4, 34, 22, 14, 24, 14]
for i,w in enumerate(widths,1): pr.column_dimensions[get_column_letter(i)].width = w
pr.row_dimensions[1].height = 28
for r in range(2, len(POLICIES)+2): pr.row_dimensions[r].height = 32

# ============================================================
# Sheet 11 - Project Plan (24 months)
# ============================================================
pp = wb.create_sheet("11 Project Plan")
hh = ["Week","Workstream","Standard","Task","Owner","Outcome"]
for c,h in enumerate(hh,1):
    cell = pp.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

PROJECT = [
    [1,"Mobilise","All 5","Charter + 5-standards SteerCo + PM appointment","CISO + CRO","Programme Charter"],
    [2,"Mobilise","All 5","Auditor / QSA / Certification Body selection","CFO","Engagement letters"],
    ["3-4","Scoping","SOC 2 + PCI","System description + CDFD","CISO + Eng","SOC 2 System Desc + PCI CDFD"],
    ["3-4","Scoping","ISO 27001","SoA initial + risk treatment plan","CISO","SoA v1.0"],
    ["3-4","Scoping","NIST CSF","Target profile + current profile assessment","CISO + Internal Audit","NIST profile v1.0"],
    ["3-4","Scoping","DPDP","Record of Processing Activities (RoPA)","DPO + DPO","RoPA v1.0"],
    ["5-6","Inventory","All 5","Asset inventory + CDE + Privacy data flow","CISO + SecOps","Inventory live"],
    ["5-8","Scoping","All 5","Joint gap assessment + tag assignment (Universal/Mostly-shared/Standard-specific/Differentiated)","CISO + Owners","Gap register"],
    ["7-12","TRA","PCI","Run TRA workshops for v4.0 mandatory requirements","CISO","TRA register"],
    ["7-12","Risk Assessment","ISO 27001 + NIST","Information security risk assessment + asset register","CISO + Risk","RA report"],
    ["7-12","DPIA","DPDP","DPIA for all significant data processing","DPO + DPO","DPIA register"],
    ["8-20","Remediation","All 5","IAM + KMS + HSM + EDR + SIEM + WAF + ASV + GRC tool","Various","Tools configured"],
    ["9-26","Policies","All 5","Write 70 documents; version control; owner ack","CISO + Owners","All 70 docs live"],
    ["10-22","Vendor","All 5","AOC collection: SOC 2 + PCI + ISO + DPDP DPA","CISO + Procurement","Vendor pack"],
    ["12-26","Pen-test","PCI + SOC 2 + ISO","Annual pen-test; segmentation test (PCI)","AppSec","PT reports"],
    ["22-32","Dry Run","SOC 2 + PCI","Mock audit + sample pulls + interviews","CISO + PM","Internal ROC + readiness"],
    ["22-32","ASV","PCI","Q1 ASV scan + re-scan any failures","SecOps","ASV pass letter"],
    ["24-36","Audit","PCI","QSA fieldwork + ROC draft","QSA + CISO","ROC v1.0"],
    ["24-36","Audit","SOC 2","Auditor fieldwork + Working Papers","Auditor + CISO","Working papers"],
    ["26-38","Audit","ISO 27001","Certification Body Stage 1 + Stage 2 audit","CB + CISO","ISO 27001 cert"],
    ["28-40","Audit","NIST CSF","Internal maturity assessment","CISO + Internal Audit","NIST profile v2.0"],
    ["30-40","Issuance","PCI","AOC + ROC + acquirer filing","CISO + CFO","AOC delivered to acquirer"],
    ["30-40","Issuance","SOC 2","Type II report issued","CISO + CFO","SOC 2 Type II"],
    ["30-40","Issuance","ISO 27001","ISO 27001 certification","CISO + CFO","ISO cert"],
    ["36-46","Audit","DPDP","Self-attestation + third-party audit","DPO + External","DPDP attestation"],
    ["40-50","Surveillance","ISO 27001","First surveillance audit (year 1)","CB + CISO","Surveillance OK"],
    ["50-60","Recertification","SOC 2","Next SOC 2 Type II cycle","CISO + Auditor","SOC 2 year 2"],
    ["52-65","Recertification","PCI","Next PCI ROC","CISO + QSA","ROC year 2"],
    ["60-72","Continuous","All 5","Quarterly dashboards + control health + GRC reporting","Compliance","Continuous"],
    ["80+","Maintenance","NIST","Annual maturity assessment","Internal Audit","Annual report"],
]
for r,row in enumerate(PROJECT,2):
    for c,v in enumerate(row,1):
        cell = pp.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==3:
            color = {
                "All 5":EMERALD, "SOC 2 + PCI":INDIGO, "PCI":AMBER, "SOC 2":INDIGO,
                "ISO 27001":GOLD, "NIST CSF":NAVY, "DPDP":AMBER,
                "ISO 27001 + NIST":INDIGO, "PCI + SOC 2 + ISO":EMERALD,
                "SOC 2 + ISO":EMERALD,
            }.get(v,"")
            if color:
                cell.fill = fill(color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [9, 22, 20, 50, 24, 28]
for i,w in enumerate(widths,1): pp.column_dimensions[get_column_letter(i)].width = w
pp.row_dimensions[1].height = 28
for r in range(2, len(PROJECT)+2): pp.row_dimensions[r].height = 32
pp.freeze_panes = "A2"

# ============================================================
# Sheet 12 - Evidence Catalog
# ============================================================
ec = wb.create_sheet("12 Evidence Catalog")
hh = ["#","Control","Standards","Frequency","Source","Tag","Pre-audit check"]
for c,h in enumerate(hh,1):
    cell = ec.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

EVIDENCE = [
    [1,"Information Security Policy + acknowledgement","All 5","Annual","LMS","UNIVERSAL","100% signed"],
    [2,"Risk register + review minutes","All 5","Quarterly","GRC","UNIVERSAL","Reviewed ≤90d"],
    [3,"BoD minutes on cyber","SOC 2 + ISO + NIST","Quarterly","Confluence","UNIVERSAL","On agenda"],
    [4,"SoA + RTP","ISO only","Annual","Confluence","STANDARD-SPECIFIC","Reviewed"],
    [5,"MFA enrolment","All 5","Quarterly","Okta/Entra","UNIVERSAL","100% CDE users"],
    [6,"IAM provisioning tickets","SOC 2 + PCI + ISO","Per request","Jira","UNIVERSAL","All approved"],
    [7,"Quarterly access reviews","SOC 2 + PCI + ISO","Quarterly","Confluence+IdP","MOSTLY-SHARED","All reviewed"],
    [8,"Card vault encryption attestation","PCI + ISO","Quarterly","AWS CloudHSM","MOSTLY-SHARED","Attestation current"],
    [9,"TLS scan A rating","SOC 2 + PCI + ISO","Quarterly","Qualys SSL","MOSTLY-SHARED","A or A+"],
    [10,"EDR coverage","All 5","Continuous","CrowdStrike","UNIVERSAL","≥99% fleet"],
    [11,"SIEM alert log + closure","SOC 2 + PCI + NIST","Continuous","Splunk","UNIVERSAL","MTTR within SLA"],
    [12,"12-month log retention","PCI","Continuous","Splunk","DIFFERENTIATED","≥12 months"],
    [13,"Quarterly ASV pass letter","PCI only","Quarterly","ASV vendor","STANDARD-SPECIFIC","Pass"],
    [14,"Internal vuln scan","SOC 2 + PCI + ISO + NIST","Quarterly","Tenable","UNIVERSAL","No criticals"],
    [15,"Pen-test report + segmentation","SOC 2 + PCI + ISO","Annual","PT vendor","UNIVERSAL","All scopes"],
    [16,"IRP + Tabletop minutes","All 5","Annual","Confluence","UNIVERSAL","Tabletop 2x/year"],
    [17,"DR drill report","All 5","Annual","Confluence + AWS","UNIVERSAL","RTO/RPO met"],
    [18,"CAB minutes + change tickets","All 5","Weekly","Jira","UNIVERSAL","All prod"],
    [19,"SP register + AOCs + ISO certs","SOC 2 + PCI + ISO","Quarterly","Compliance","UNIVERSAL","AOC current"],
    [20,"MSA with PCI ack clause","PCI only","Per vendor","Legal","STANDARD-SPECIFIC","Signed"],
    [21,"DPA with data principals","DPDP only","Per vendor","Legal","STANDARD-SPECIFIC","Signed"],
    [22,"Consented receipt + withdrawal","DPDP only","Continuous","Backend","STANDARD-SPECIFIC","100% captured"],
    [23,"Data Principal Rights workflow","DPDP only","Per request","Support","STANDARD-SPECIFIC","SLAs met"],
    [24,"RoPA","DPDP + ISO","Quarterly","DPO tool","MOSTLY-SHARED","Quarterly refresh"],
    [25,"Privacy notice version","DPDP + SOC 2 + ISO","Annual","Web + backend","STANDARD-SPECIFIC","Live"],
    [26,"Privacy training","All 5","Annual","LMS","UNIVERSAL","≥95% completion"],
    [27,"Phishing simulation results","SOC 2 + PCI 5.4.1 + NIST","Quarterly","KnowBe4","MOSTLY-SHARED","≥-50% phish-prone"],
    [28,"DPIA report","DPDP only","Per new processing","DPO","STANDARD-SPECIFIC","Completed"],
    [29,"TRA register","PCI 4.0 only","Annual","GRC","STANDARD-SPECIFIC","12 TRAs"],
    [30,"Script inventory + SRI","PCI 6.4.3","Weekly","AppSec","STANDARD-SPECIFIC","100% authorised"],
    [31,"POI monthly inspection","PCI 9.5","Monthly","Operations","STANDARD-SPECIFIC","All completed"],
    [32,"3DS2 configuration","PCI 3DS2","Continuous","Payments","STANDARD-SPECIFIC","Enabled"],
    [33,"SBI test","not standard","Annual","FPSB","STANDARD-SPECIFIC","(Optional SBI capability)"],
    [34,"BCP RTO/RPO profiling","All 5","Annual","BIA","UNIVERSAL","Documented"],
    [35,"BCP tabletop","All 5","Bi-annual","Confluence","UNIVERSAL","2x/year"],
    [36,"Cyber insurance policy","SOC 2 only","Annual","CFO","STANDARD-SPECIFIC","In force"],
    [37,"Information security metrics dashboard","All 5","Continuous","GRC","UNIVERSAL","Live"],
    [38,"Audit committee report","All 5","Quarterly","CISO + CRO","UNIVERSAL","Submitted"],
    [39,"SAMA Gap analysis","SOC 2 + ISO + NIST","Annual","CISO","MOSTLY-SHARED","Annual"],
    [40,"Vendor incident coordination","All 5","Per incident","Compliance","UNIVERSAL","Acknowledgement <24h"],
    [41,"Crypto agility assessment","NIST only","Annual","Crypto Lead","STANDARD-SPECIFIC","PQC roadmap"],
    [42,"AI/LLM risk assessment","DPDP + NIST","Annual","Product Security","MOSTLY-SHARED","TRA filed"],
    [43,"Lab Inventory / SBOM","PCI 6.3.3","Monthly","AppSec","STANDARD-SPECIFIC","Per release"],
    [44,"Annual compliance schedule","All 5","Annual","Compliance","UNIVERSAL","Cover all 5"],
    [45,"RBI Cyber Security Reporting Register","RBI only","Continuous","Compliance","STANDARD-SPECIFIC","SLA met"],
    [46,"RBI Master Direction Outsourcing Register","RBI only","Continuous","Compliance","STANDARD-SPECIFIC","Reviewed"],
    [47,"RBI Tokenisation Compliance","RBI only","Continuous","CTO","STANDARD-SPECIFIC","CoF tokenised"],
    [48,"Certificate inventory","All 5","Continuous","Crypto Lead","UNIVERSAL","Renewed"],
    [49,"Customer comms test (status page)","SOC 2 + ISO","Continuous","StatusPage","MOSTLY-SHARED","Verified"],
    [50,"Audit log review aging","SOC 2 + PCI","Continuous","SOC","UNIVERSAL","<24h"],
]
for r,row in enumerate(EVIDENCE,2):
    for c,v in enumerate(row,1):
        cell = ec.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==6:
            color = {
                "UNIVERSAL":EMERALD, "MOSTLY-SHARED": INDIGO, 
                "STANDARD-SPECIFIC":AMBER, "DIFFERENTIATED":RED,
            }.get(v,"")
            if color:
                cell.fill = fill(color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [4, 32, 28, 14, 22, 18, 30]
for i,w in enumerate(widths,1): ec.column_dimensions[get_column_letter(i)].width = w
ec.row_dimensions[1].height = 28
for r in range(2, len(EVIDENCE)+2): ec.row_dimensions[r].height = 32

# ============================================================
# Sheet 13 - Vendor Register
# ============================================================
vr = wb.create_sheet("13 Vendors")
hh = ["#","Vendor","Service","SOC 2","PCI","ISO","DPDP","Risk","Owner"]
for c,h in enumerate(hh,1):
    cell = vr.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

VENDORS = [
    [1,"AWS India","Hosting/KMS/CloudHSM","Yes","Yes (SP L1)","Yes","DPA","Low","Cloud Lead"],
    [2,"Okta","SSO + WebAuthn","Yes","n/a","Yes","DPA","Low","IAM Lead"],
    [3,"CyberArk","PAM","Yes","n/a","Yes","n/a","Low","IAM Lead"],
    [4,"CrowdStrike","EDR","Yes","n/a","Yes","n/a","Low","SecOps"],
    [5,"Splunk","SIEM","Yes","Yes (PCI attested)","Yes","n/a","Low","SOC Lead"],
    [6,"Tenable","Vulnerability scanner","Yes","Yes","Yes","n/a","Low","SecOps"],
    [7,"ASV (Trustwave / CyberOxide)","Quarterly ASV","n/a","Yes","n/a","n/a","Low","SecOps"],
    [8,"Visa Token Service (VTS)","Tokenisation","n/a","Yes","n/a","n/a","Low","CTO"],
    [9,"Mastercard MDES","Tokenisation","n/a","Yes","n/a","n/a","Low","CTO"],
    [10,"RuPay Token Vault","Tokenisation (IN)","n/a","Yes","n/a","n/a","Low","CTO"],
    [11,"Razorpay / Cashfree / Adyen","Payment gateway","Yes","Yes","Yes","DPA","Low","CTO"],
    [12,"Pine Labs / Innoviti","P2PE POS","n/a","Yes (P2PE)","n/a","n/a","Low","Operations"],
    [13,"AuthBridge / HireRight","BGV","Yes","n/a","n/a","n/a","Low","HR"],
    [14,"KnowBe4 / Hoxhunt","Phishing sim","Yes","n/a","Yes","n/a","Low","CISO"],
    [15,"Bajaj Finance / ICICI","Co-issuer bank","n/a","Yes","Yes","DPA","Med","CTO"],
    [16,"Signzy / Perfios","eKYC V-CIP","Yes","No (no CHD)","Yes","n/a","Med","Product"],
    [17,"Pen-test vendor (NCC / NetSPI)","Pen-test","Yes","Yes","Yes","n/a","Low","AppSec"],
    [18,"Collection agencies","Field collection","Yes","No (no CHD)","n/a","DPA","High","Compliance"],
    [19,"Big-4 Audit firm","SOC 2 + ISO","Yes","Yes","Yes","n/a","Low","CFO"],
    [20,"Tier-1 QSA firm","PCI ROC","Yes","Yes","Yes","n/a","Low","CFO + CISO"],
    [21,"CIBIL / Experian / Equifax","Credit bureau","Yes","n/a","n/a","DPA","Med","Credit Risk"],
    [22,"AI / LLM provider","Provider LLM","Yes","n/a","Yes","DPA","Med","Product"],
    [23,"Translation vendor","Localization","Yes","n/a","n/a","DPA","Low","Marketing"],
    [24,"Backup vendor (NetApp)","Backup + DR","Yes","Yes","Yes","n/a","Low","IT"],
    [25,"AV / Anti-malware","EDR alternative","Yes","Yes","Yes","n/a","Low","IT"],
]
for r,row in enumerate(VENDORS,2):
    for c,v in enumerate(row,1):
        cell = vr.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==8:
            if v=="High":
                cell.fill = fill(RED); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif v=="Med":
                cell.fill = fill(AMBER); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [4, 22, 22, 10, 10, 10, 10, 9, 14]
for i,w in enumerate(widths,1): vr.column_dimensions[get_column_letter(i)].width = w
vr.row_dimensions[1].height = 28
for r in range(2, len(VENDORS)+2): vr.row_dimensions[r].height = 32

# ============================================================
# Sheet 14 - Multi-Audit Q&A
# ============================================================
qa = wb.create_sheet("14 Audit Trail Q&A")
hh = ["#","Category","Question","Answer","Source tab"]
for c,h in enumerate(hh,1):
    cell = qa.cell(row=1,column=c,value=h); cell.fill=fill(NAVY); cell.font=head_font(); cell.alignment=cell_wrap(); cell.border=BORDER

QUESTIONS = [
    [1,"Common","How do 5 standards run together?","Same SteerCo, same evidence repo, dedicated tag per control. Triple-tagged evidence: SOC 2 + PCI + ISO + NIST + DPDP refs in each file's YAML.","09 + Tab 03"],
    [2,"Common","Cost saving vs five separate programmes?","~40% engineer-hour reduction; audit fees from ₹5+ crore/y to ₹3+ crore/y joint headline cost; 8 dashboards collapse to 1.","Joint PDF + 11 Plan"],
    [3,"Common","How many controls total?","~100 deduped controls from >250 raw requirements across 5 standards. ~50% dedupe rate achievable.","09 Checklist"],
    [4,"Common","Biggest engineering hour saver?","Identity & Access: single Okta + WebAuthn replaces 5 standards' onboarding workflows. Saves ~28% total.""09 Task 9-15"],
    [5,"SOC 2","TSC set chosen","Security, Availability, Confidentiality, Processing Integrity. Privacy deferred to 2026 cycle.","SOC 2 PDF"],
    [6,"SOC 2","Type II timing","12-month window aligned with RBI fiscal year (Apr-Mar) or Jan-Dec per board preference.","SOC 2 PDF"],
    [7,"PCI","Service provider Level 1?","Yes; 2.6M cards in vault; ROC by Tier-1 QSA; quarterly ASV by Trustwave India.","PCI PDF"],
    [8,"PCI","Tokenisation per RBI?","CoF all tokenised via VTS/MDES/R-T-V; >98% redaction on store. Reduces effective PCI scope by ~28%.","PCI PDF"],
    [9,"PCI","TRA register — how many?","12 TRAs in active register; reviewed annually + on material change.","PCI PDF + 12"],
    [10,"DPDP","SDF obligations","Significant Data Fiduciary provisions triggered >vol> annual retention; DPIA needed for >large scale special category.","DPDP PDF + 04"],
    [11,"DPDP","72-hour breach","Yes - 72h to data principals; 2h to RBI for severe incidents; parallel clocks governed via single IRP.","DPDP PDF + 08"],
    [12,"DPDP","Cross-border lock","Region lock to ap-south-1 + ap-southeast-1; no transfer to non-notified countries.","04 + Sheet 4"],
    [13,"ISO 27001","SoA kept?","Yes; A.5 increments across all 93 Annex A controls. Aerated annual.","05 + 10"],
    [14,"ISO 27001","Internal audit programme","Yes; 3-yr cycle; cover all 5 standards' clauses.","05 + 14"],
    [15,"NIST CSF","2.0 changes","Added Govern function; expanded DE.AE/DE.CM; added GV.SC; mandates integrated risk.","06"],
    [16,"NIST CSF","Target profile?","Tier 3 → 4 trajectory; biannual maturity assessment.","06 + 11"],
    [17,"Cross","Which framework unifies?","NIST CSF 2.0 since RBI IT Framework 2023 maps closely to NIST 800-53 + CSF","06 + 02"],
    [18,"Cross","Single evidence index","Each file has YAML header 'tsp_id','pci_req','iso','nist','dpdp' fields. GRC tool indexes.","Joint PDF"],
    [19,"Cross","Joint calendar","PC Sep-Oct / SOC 2 Jan-Feb / ISO 27001 Stage 1 in Q1, Stage 2 in Q2 / NIST in Q3 / DPDP in Q4.","Joint PDF + 11"],
    [20,"Cross","AOC expiry alerts","Auto alert 60/30/15 days pre-expiry. SP register refreshed quarterly.","13 + 09"],
    [21,"NBFC","RBI Cyber Committee met?","Quarterly; chaired by MD/CEO; minutes circulated.","Sheet 02 + 10"],
    [22,"NBFC","DPDP SDF obligations met?","DPIA again for any new product launch (e.g., co-lending card product).","04 + 09"],
    [23,"NBFC","CIBIL data classification","Confidential tier; last-4 only in non-CDE stores; full pull logs in vault.","09 + 13"],
    [24,"NBFC","Field collection agency risk","High risk; AOCs received + DPA signed + TRA done + quarterly audit rights.","13"],
    [25,"Cost","Auditor fees combined","SOC 2 ₹12-25 lakh + PCI ROC ₹18-50 lakh + ISO surveillance ₹10-25 lakh + NIST maturity ₹5-10 lakh = ₹45-110 lakh/yr","Joint PDF"],
    [26,"Cost","Engineer-hour saving","Approx 2,000 engineer-hours saved annually vs 5 separate programmes.","Joint PDF"],
    [27,"Cost","GRC tool annual cost","Vanta ₹35-65 lakh/yr; Drata ₹30-55 lakh; Tugboat cheaper.","Joint PDF"],
    [28,"Risk","Biggest implementation risk?","Parallel deadlines; AOC expiry slippage; DPDP penalty exposure.","Risk Plan"],
    [29,"Risk","Mitigation","Quarterly CISO review; auto calendar; quarterly board report.","Risk Plan"],
    [30,"Tool","Single evidence repository","S3 versioning + locking; manifests per TSC + PCI Req + ISO Annex + NIST Subcategory + DPDP §.","12 evidence catalog"],
    [31,"Tool","Sample pull cadence","30-40 per control per standard; cross-tagging in evidence file metadata.","12"],
    [32,"Tool","Cross framework dashboards","Vanta + custom Power BI / Tableau dashboards; weekly refresh.","Joint PDF"],
    [33,"Tool","Pen-test vendor","Single NCC Group / NetSPI covers SOC 2 + PCI + ISO + NIST tests in one statement of work.","Joint PDF"],
    [34,"Maintenance","Annual work","Q1 SOC 2 renewal; Q2 ISO surveillance; Q3 PCI ROC start; Q4 NIST maturity.","11"],
    [35,"Maintenance","Quarterly work","Quarterly CCO committee + AOC expiry check + control health pull.","11"],
    [36,"Vendor","Vendor pyramid","AWS + Azure (T-II SaaS) + cyber tools (T-II SaaS); token + payment ecosystem (T-II PCI).","13"],
    [37,"Vendor","Manage 25+ vendors","Quarterly CISO review; automated calendar reminders; annual onsite review for critical 5.","13"],
    [38,"Vendor","Risks of SP consolidation","Single source of truth risk on centralised services (e.g., Okta).","13"],
    [39,"Process","SteerCo meetings","Monthly operational + Quarterly SteerCo with Board Cyber Committee.","Joint PDF"],
    [40,"Process","Auditor rotation","Rotate auditor + QSA + CB every 5 years for independence.","Joint PDF"],
]
for r,row in enumerate(QUESTIONS,2):
    for c,v in enumerate(row,1):
        cell = qa.cell(row=r,column=c,value=v); cell.alignment=cell_wrap(); cell.border=BORDER
        if c==2:
            color = {
                "Common":EMERALD, "SOC 2":INDIGO, "PCI":AMBER, "DPDP":GOLD,
                "ISO 27001":GOLD, "NIST CSF":GOLD, "Cross":INDIGO,
                "NBFC":GOLD, "Cost":EMERALD, "Risk":RED, "Tool":INDIGO,
                "Maintenance":EMERALD, "Vendor":EMERALD, "Process":SLATE,
            }.get(v,"")
            if color:
                cell.fill = fill(color); cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
widths = [4, 14, 38, 70, 22]
for i,w in enumerate(widths,1): qa.column_dimensions[get_column_letter(i)].width = w
qa.row_dimensions[1].height = 28
for r in range(2, len(QUESTIONS)+2): qa.row_dimensions[r].height = 36

wb.save(OUT)
print(f"Wrote {OUT}")
size = os.path.getsize(OUT)
print(f"Size: {size:,} bytes")
